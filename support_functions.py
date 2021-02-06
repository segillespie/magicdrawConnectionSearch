def XLSX_to_DF(directory, filename, sheet_name, skip = None):
    #after exporting table from magicdraw it appears to be helpful to copy and 
    #repaste values only.  this removes the formatting and allows pandas to more
    #clearly read the cells, especially cells that are delimited by a new line '\n'
    #as is the case for the connected list.
    
    import pandas as pd # storing output as pandas data frame
    import os #required to navigate to appropriate directory
    import math
    import numpy as np
    
    #inputs for to be function
    #directory = '/Users/gillespie/Desktop'
    #filename = 'Model_Output_Test.xlsx'
    #sheet_name = 'Sheet1'
    
    os.chdir(directory) #change to current directory
    
    df = pd.read_excel(filename, sheetname = sheet_name, skiprows = skip) #read .xlsx file and convert to data frame
    
    
    ###############################################################################
    # Clean the Data
    
    # delete column consisting solely of NaN
    for col in df.columns:
        c = True #initially assume the column is empty
        for ind in df.index: #check each index in the column 
            if not((type(df[col][ind]) in [float, np.float64]) and math.isnan(df[col][ind])): #if the entry at column / index is anything but NaN, the column is non-empty, return false and quit checking that index
                c = False
                break
        if c: #if every index is empty, the column is empty, delete the column
            del df[col]
    
    # delete index consisting soleley of NaN
    # same algorithm as for columns
    for ind in df.index:
        i = True
        for col in df.columns:
            if not((type(df[col][ind]) in [float, np.float64]) and math.isnan(df[col][ind])):
                i = False
                break
        if i:
            df = df.drop(ind)
            
    for col in df.columns: #iterate through each column in the dataframe   
        #check if there is a string in the column (can't just check first element as it may be a NaN of type float)
        if df[col].dtype == 'O' and df[col].str.contains('\n').any(): 
            # iterate through columns with strings that need to be replaced by lists
            newlist = []
            for c in df[col]:
                # pandas reads empty cells as NaN of type float, and other cells as strings
                # check to see if the cell is a float of type NaN and then assign it as no
                # connections, i.e. an empty list
                if type(c) == str:
                    newlist.append(c.split('\n'))
                elif type(c) == float and math.isnan(c):
                    newlist.append([])
                elif type(c) == float or type(c) == int:
                    newlist.append([c])
                else: #otherwise cell will be 
                    newlist.append(c.split('\n'))
            df[col + '_split'] = newlist
            del df[col]
    
    return df
    
    #######################################################
    
#defines support functions for search
# all_connectors_attached_to_element 

def all_connectors_attached_to_element(element, df_connector, df_part, df_port):
    #takes a given element name, searches the appropriate file to see if it is there and then returns a list with all connectors by ID attached to that element. If there are none, it returns an empty list.
    #search connectors on part
    from support_functions import element_idx
    elements_to_search = []
    elements_to_search.append(element)
    ref_elements = 1.2 #just need a dummy float
    [e_idx, e_type] = element_idx(element, df_port, df_part, df_connector)
    #find all parts that are bound to the given element and add them to the list of elements to search
    if e_type == 'part': 
        if 'Binding Connector (Ref Property) ElmID_split' in df_part.columns: key = 'Binding Connector (Ref Property) ElmID_split'
        else: key = 'Binding Connector (Ref Property) ElmID'
        ref_elements = df_part[key][e_idx]
    if type(ref_elements) == str: 
        elements_to_search.append(ref_elements)
    elif type(ref_elements) == list:
        for re in ref_elements: elements_to_search.append(re)
    
    #find all connectors that have a 'part with port' equal to the element or the referred element(s)
    s = df_connector['Part with Port for Connector Ends ElmID_split']
    connectors = []
    for ets in elements_to_search:
        idx = list(s[s.apply(lambda i: ets in i)].index.values) #no idea how this works, but it does.  thank you stackoverflow https://stackoverflow.com/questions/47439796/find-elements-of-a--pandas-series-of-lists-that-contain-a-specific-value
        for i in idx: connectors.append(df_connector['Element ID'][i])
    return connectors


#returns all connectors that have an end at this port
def all_connectors_attached_to_port(current_role, df_connector):
    s = df_connector['ElmID of Role of Connector Ends (Port)_split']
    idx = list(s[s.apply(lambda i: current_role in i)].index.values) #no idea how this works, but it does.  thank you stackoverflow https://stackoverflow.com/questions/47439796/find-elements-of-a--pandas-series-of-lists-that-contain-a-specific-value
    connectors = []
    for i in idx: connectors.append(df_connector['Element ID'][i])
    return [connectors, idx]



#defines support functions for search
# all_connectors_attached_to_element 

def all_connectors_owned_by_element(e_type, df_connector):
    #takes a given element name, searches the appropriate file to see if it is there and then returns a list with all connectors by ID attached to that element. If there are none, it returns an empty list.
    #search connectors on part
    
    s = df_connector['Owner']
    idx = list(s[s.apply(lambda i: e_type == i)].index.values) #no idea how this works, but it does.  thank you stackoverflow https://stackoverflow.com/questions/47439796/find-elements-of-a--pandas-series-of-lists-that-contain-a-specific-value
    connectors = []
    for i in idx: connectors.append(df_connector['Element ID'][i])
    return connectors



def element_idx(element, df_port, df_part, df_connector):
    if element in df_connector['Element ID'].values:
        e_type = 'connector'
        e_idx = df_connector[df_connector['Element ID'] == element].index.tolist()
    elif element in df_part['Elm ID'].values:
        e_type = 'part'
        e_idx = df_part[df_part['Elm ID'] == element].index.tolist()
    elif element in df_port['Elm ID'].values:
        e_type = 'port'
        e_idx = df_port[df_port['Elm ID'] == element].index.tolist()
    else:
        return [-1, 'no type']
    return[e_idx[0], e_type]
    



#pb = port_between(path[idx], path[idx+1]) #returns [port_elm_id, port_idx]; if no port, returns []
def port_between(elm1, elm2, df_port, df_part, df_connector):
    from support_functions import element_idx
    e1 =element_idx(elm1, df_port, df_part, df_connector)
    e2 =element_idx(elm2, df_port, df_part, df_connector)
    
    ############################################################################
    #situation 1, the inputs are a part and connector
    if [e1[1], e2[1]] in [['part', 'connector'], ['connector', 'part']]:
        # ID which part is a connector and which is the part
        if e1[1] == 'connector':
            connector = elm1
            c = e1
            part = elm2
            p = e2
        else:
            connector = elm2
            c = e2
            part = elm1
            p = e1
        #find the connector roles 
        connector_roles = list(df_connector['ElmID of Role of Connector Ends (Port)_split'][c[0]])
        #check each connector role to see if 1) the part is the role - in which case there is no port between, or if the role is a port owned by the type of the part 
        #(note this could possibly get screwed up in a wierd instance in which a connector is attached to two parts of the same type, but at different ports on the part....
        for r in connector_roles:
            if r == part:
                return []
            elif (r in df_port['Elm ID'].values):
                r_idx = element_idx(r, df_port, df_part, df_connector)
                if df_port['Owner'][r_idx[0]] == df_part['Type'][p[0]]:
                    return [r, r_idx[0]]
            else:
                return []   
    ############################################################################
    #situation 2, the inputs are a both connectors, find their shared role
    elif [e1[1], e2[1]] == ['connector', 'connector']:
        e1roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][e1[0]]
        e2roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][e2[0]]
        shared_role = list(set(e1roles).intersection(set(e2roles)))        
        if (len(shared_role) == 1) and (shared_role[0] in df_port['Elm ID'].values):
            return [shared_role[0], element_idx(shared_role[0], df_port, df_part, df_connector)[0]]
        else:
            return []
    ############################################################################
    #situation 3, sit. 1 or 2 don't apply, return empty
    else:
        return []
    




# returns the element ID of the port that is on that part
def port_on_part(c_idx, e_idx, df_connector, df_part, df_port):
    connector_ports = df_connector['ElmID of Role of Connector Ends (Port)_split'][c_idx[0]] # returns the ports that are at the ends of the given connector
    for cp in connector_ports: #for each elmID of the ports listed
        p_idx = df_port[df_port['Elm ID'] == cp].index.tolist() #find the index of that port within df_port
        if df_port['Owner'][p_idx[0]] == df_part['Type'][e_idx[0]]: #check to see if the owner of that port is equal to the type of that part
            return [cp] #if this is true, then that means that we're looking at the appropriate part - #note th
    #if neither of the ports on the connector are owned by the type for the part that we are looking at, then there is no port to check.
    return []




def port_to_connectors(Current_Connector, Current_Port, previous, df_connector, df_part, df_port):
    
    from support_functions import all_connectors_attached_to_port # require this function
    
    #get the current connector's index in df_connector and it's owner
    Current_Connector_IDX = df_connector[df_connector['Element ID'] == Current_Connector].index.tolist()
    Current_Connector_Owner = df_connector['Owner'][Current_Connector_IDX[0]]
    
    #get the current port's index, owner, and all connectors attached to the port by index and name
    Current_Port_IDX = df_port[df_port['Elm ID'] == Current_Port].index.tolist()
    Current_Port_Owner = df_port['Owner'][Current_Port_IDX[0]]
    All_Connectors_Attached_To_Current_Port_Name = list(all_connectors_attached_to_port(Current_Port, df_connector)[0]) #edited to create copy, not just pointers
    All_Connectors_Attached_To_Current_Port_IDX = list(all_connectors_attached_to_port(Current_Port, df_connector)[1]) #edited to create a copy, not just pointers

    
    # there are two cases to consider, one in which 
    if Current_Port_Owner == Current_Connector_Owner: # if the port we're looking at and the connector we're viewing, then we're going "out"
        i = len(All_Connectors_Attached_To_Current_Port_Name) - 1
        while i >=0:
            if (All_Connectors_Attached_To_Current_Port_Name[i] in previous) or (df_connector['Owner'][All_Connectors_Attached_To_Current_Port_IDX[i]] == Current_Port_Owner): 
                #don't look at connectors that have already been checked, don't look at connectors that are "inside" the block
                del All_Connectors_Attached_To_Current_Port_IDX[i]
                del All_Connectors_Attached_To_Current_Port_Name[i]
            i = i-1
        return All_Connectors_Attached_To_Current_Port_Name
            
    else: # implies current port owner is not equal to current connector, so we're looking "in"
        #find all connectors attached to this port
            i = len(All_Connectors_Attached_To_Current_Port_Name) - 1
            while i >= 0:
                if (All_Connectors_Attached_To_Current_Port_Name[i] in previous) or df_connector['Owner'][All_Connectors_Attached_To_Current_Port_IDX[i]] != Current_Port_Owner:
                    del All_Connectors_Attached_To_Current_Port_IDX[i]
                    del All_Connectors_Attached_To_Current_Port_Name[i]
                i = i-1
            return All_Connectors_Attached_To_Current_Port_Name #this will be a list of all connectors that 1) have not already been searched & 2) are "in" the block of the current port
                
    
    #partproperty = '_18_5_2_8bc025c_1507816075657_952311_113567' #JC Left Shoulder
    #connector = '_18_5_2_5f601d8_1508168020605_106519_274481' #e-stop to e-stop left shoulderJC Left shoulder
    #port = '_18_5_2_5f601d8_1508168020606_208085_274484' # e stop out of left shoulder sensing & actuation group
    

def list_or_set_to_string(ls, empty_message = 'Empty', delim = ', '):
    if len(ls) == 0:
        return empty_message
    elif len(ls) == 1:
        return str(list(ls)[0])
    elif len(ls) > 1:
        newstring = ''
        for item in list(ls):
            newstring += str(item) + delim
        return newstring[0:(len(newstring) - len(delim))]
    else:
        return 'Error. list_or_set_to_string.'    



def context_in_english(context, df_connector, df_part, df_port):
    from support_functions import element_idx
    if not context:
        return []
    else:
        result = []
        for c in context:
            [c_idx, c_type] = element_idx(c, df_port, df_part, df_connector)
            if c_type == 'part':
                if str(df_part['Name'][c_idx]) != 'nan':
                    result.append(df_part['Name'][c_idx])
                elif str(df_part['Name'][c_idx]) == 'nan' and str(df_part['Type'][c_idx]) != 'nan':
                    result.append('Unnamed ' + df_part['Type'][c_idx])
                else: # str(df_part['Name'][c_idx]) == 'nan' and str(df_part['Type'][c_idx]) != 'nan':
                    result.append('Unnamed / untyped: ' + c)

            else:
                result.append('Context ID not a part in DataFrame: ' + c)
    
    return result
    
################################################################################    
def idWireHarnessOnPath(path, df_connector, df_part, df_port):
    # this function takes a path composed of parts, connectors, and ports
    # searches through each member of the path, and IDs all elements allocated
    # from each connector in the path
    # it then searches each element allocated from that connector and ensures 
    # that it is 1) a part property and 2) a wire harness
    # it then returns the set of all wire harnesses allocated from this path
    from support_functions import element_idx
    
    wiringHarnesses = []
    for p in path:
        [e_idx, e_type] = element_idx(p, df_port, df_part, df_connector)
        
        
        if e_type == 'connector':
            allocations = df_connector['Allocated From ElmID'][e_idx]
        
        if type(allocations) == str: wiringHarnesses.append(allocations)
        elif type(allocations) == list:
            for a in allocations:
                if type(a) == str: wiringHarnesses.append(a)
    
    wh_keep = []
    
    for wh in wiringHarnesses:
        [wh_idx, wh_type] = element_idx(wh, df_port, df_part, df_connector)
        if wh_type == 'part':
            if df_part['Is Wire Harness'][wh_idx]: wh_keep.append(wh)
    
    return wh_keep
    
    
def ref_to_part(element, df_part):
    #takes a reference property element ID and returns the element ID(s) of part properties that have a binding connector with it
    if 'Binding Connector (Ref Property) ElmID_split' in df_part.columns: key = 'Binding Connector (Ref Property) ElmID_split'
    else: key = 'Binding Connector (Ref Property) ElmID'
    s = df_part[key]
    idx = list(s[s.apply(lambda i: element == i)].index.values) #no idea how this works, but it does.  thank you stackoverflow https://stackoverflow.com/questions/47439796/find-elements-of-a--pandas-series-of-lists-that-contain-a-specific-value
    print(idx)
    referred_parts = []
    for i in idx: referred_parts.append(df_part['Elm ID'][i])
    return referred_parts
    
# this identifies any wire harness allocated to the connectors in a given path    
def allocatedWireHarness(path, df_connector, df_part, df_port):
    from support_functions import element_idx
    
    WH = set()
    
    for e in path:
        [e_idx, e_type] = element_idx(e, df_port, df_part, df_connector)
        if e_type == 'connector':
            if 'Allocated From ElmID' in df_connector.columns:
                allocations = df_connector['Allocated From ElmID'][e_idx]
            elif 'Allocated From ElmID_split' in df_connector.columns:
                allocations = df_connector['Allocated From ElmID_split'][e_idx]
            if type(allocations) == str:
                [a_idx, a_type] = element_idx(allocations, df_port, df_part, df_connector)
                if a_type == 'part':
                    if df_part['Is Wire Harness'][a_idx] and not df_part['Is Abstract'][a_idx]: WH.add(allocations)
            elif type(allocations) == list:
                for a in allocations:
                    [a_idx, a_type] = element_idx(a, df_port, df_part, df_connector)
                    if a_type == 'part':
                        if df_part['Is Wire Harness'][a_idx] and not df_part['Is Abstract'][a_idx]: WH.add(a)                
    return WH
    
def getKind(element, df_part):
    if element not in df_part['Elm ID'].values: return []
    e_idx = df_part[df_part['Elm ID'] == element].index.tolist()
    kind = []
    
    if df_part['Is Wire Harness'][e_idx[0]]: kind.append('Wire Harness')
    if df_part['Is Software'][e_idx[0]]: kind.append('Software')
    if df_part['Is Exoskeleton Structure'][e_idx[0]]: kind.append('Exoskeleton Structure')

    #if df_part['Is Sensor'][e_idx[0]]: kind.append('Sensor')
    #if df_part['Is Actuator'][e_idx[0]]: kind.append('Actuator')
    #if df_part['Is Control Board'][e_idx[0]]: kind.append('Control Board')
    return kind
    
    
#Note this only works if the assemblies are uniquely identified / used as parts.
# This will be in accurate for parts that are 
def getAssyTier(element, df_part):
    if element not in df_part['Elm ID'].values: return ['Error: Input element not in df_part']

    owner = df_part['Owner'][df_part[df_part['Elm ID'] == element].index.tolist()[0]]
    inversepaths = [owner]
    
    while owner != 'FULL TALOS Assembly':
        idx = df_part[df_part['Type'] == owner].index.tolist()
        if not idx: break
        nextowner = df_part['Owner'][idx[0]]
        inversepaths.append(nextowner)
        owner = nextowner

    inversepaths.reverse()
    if inversepaths[0] != 'FULL TALOS Assembly': return ['Error: Path ends before FULL TALOS Assembly']
    
    return inversepaths






#this searches a given assembly to find all sub assemblies down to CIs.

#the result will look like: 
#[[ Super-Assembly, [Assembly1, [Sub-Assembly1.1, [CI1.1.1, CI1.1.2, ...]], [Sub-Assembly1.2, [CI1.2.1, CI1.2.2, ...]], ...], [Assembly2, ...]]]

# search requires df_part as defined in DF_Definition.py
# you can search either using a part element ID or the type (i.e. Block Name), either one as a string.
# i.e. assysearch('UEX-L Assembly', df_part) or assysearch('_12345...', df_part) 
# note in the latter search, it is the element ID of a part property, not a block

# input element requires either the element ID number as a string, e.g. '_12345'
# or a part property type name as a string, e.g. 'Joint Controller'
# or a part property owner as a string, e.g. 'Left Ankle Sensing and Actuation Group'
# this will have to match what the strings are precisely in MagicDraw

# input df_part is the dataframe for the part properties that is defined using DF_Definition.py
# input withQN defaults to True, but may be set to False if you don't want the full qualified name output
  
def assysearch(element, df_part, withQN = True):
        ########################################################################
        # this section of code identifies the input element's type and also qualified name
        # this allows the user to input a part property, Type, or Owner to start the search
        if element not in df_part['Elm ID'].values:
            if element in df_part['Type'].values:
                element_type = element
                el_QualifiedName = 'Qualified Name Unavailable'
            elif element in df_part['Owner'].values:
                element_type = element
                el_QualifiedName = 'Qualified Name Unavailable'
            else:
                return [['Error: ' + element + ' not a part property.']]
        else:
            e_idx = df_part[df_part['Elm ID'] == element].index.tolist()
            element_type = df_part['Type'][e_idx[0]]
            el_QualifiedName = df_part['Qualified Name'][e_idx[0]]
            
        ########################################################################
        # just visual output to see what is being searched
        print('Searching: ' + str(element_type))
        
        # this identifies all part properties in the model that are owned by the input element's type
        # e.g., if the input element was JC-Hip : Joint Controller
        # this finds all part properties that are owned by Joint Controller
        s = df_part['Owner']
        all_parts_owned_by_element = list(s[s.apply(lambda i: element_type == i)].index.values)
        print('All parts owned by ' + str(element_type) + ':')
        print(all_parts_owned_by_element)
        
        # if the element has no owned part properties, return the desired output (with or without qualified name)
        # note that the output is a a list containing a single list - this allows us to write the result recursively later
        if not all_parts_owned_by_element: # if element has no part properties
            if withQN:
                return [[str(element_type) + ' (' + str(el_QualifiedName) + ')', 'Last element has no part properties, not a CI.']]
            else:
                return [[str(element_type), 'Last element has no part properties, not a CI.']]
        
        # if the element does have part properties, we start the final result with the element's type & qualified name    
        if withQN:
            final = [str(element_type) + ' (' + str(el_QualifiedName) + ')']
        else:
            final = [str(element_type)]

        # iterate through each part owned by the element, check to see if it is a CI or not.  If it is a CI, then stop the search, otherwise continue the search
        for p in all_parts_owned_by_element:
            print(p)
            if df_part['CI Indicator (System Context)'][p] in ['CI', 'Ci', 'ci', 'Yes', 'yes', 'YES', 'y', 'Y']:
                if withQN:
                    final.append([df_part['Type'][p] + ' (' + str(df_part['Qualified Name'][p]) +')', 'Last element, is a CI.'])
                else:
                    final.append([df_part['Type'][p], 'Last element, is a CI.'])
            else:
                # search the part property for its owned parts
                result = assysearch(df_part['Elm ID'][p], df_part, withQN)
                # add the various results to the final list [element [part1 ], [part 2], ...]
                for r in result:
                    final.append(r)
        return [final]




def write_assy_paths(assembly, current_row, path, directory, filename, sheetname):
    import os
    import openpyxl
    import numpy as np
    

    os.chdir(directory)
    # Column 1 is element, Column 2 is End Message, Column 3 onward are Tier 1, 2, 3, ...
    
    
    for idx, sub in enumerate(assembly):
        print('Trying:')
        print(assembly)
        print('Index: ' + str(idx))
        print('Sub: ' + str(sub))
        print('Previous:')
        print(path)
        
        if idx == 0:
            path.append(sub)
            
        elif type(sub) == str:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.get_sheet_by_name(sheetname)  #get_sheet_by_name(sheetname)
            for i in np.arange(len(path)-1):
                sheet.cell(row = current_row, column = i+3).value = path[i]
            sheet.cell(row = current_row, column = 1).value = path[len(path)-1]
            sheet.cell(row = current_row, column = 2).value = sub
            current_row += 1
            wb.save(filename)
            del path[len(path)-1]
            return [current_row, path]
        else:
            [current_row, path] = write_assy_paths(sub, current_row, path, directory, filename, sheetname)
    
    del path[len(path) - 1]
    return [current_row, path]
    
    
    
    
def write_assy_to_excel(assembly, current_row, current_column, directory, filename, sheetname):
    import os
    import openpyxl
    
    #directory = '/Users/gillespie/Desktop/MagicDraw_Query/Updated_Model_Output'
    #filename = 'assembly_spreadsheet_972.xlsx'
    #sheetname = 'Without Qualified Name' 
    os.chdir(directory)
    current_row += 1
    start_column = current_column
    
    
    for idx, sub in enumerate(assembly):
        print(sub)
        if idx == 0:
            print('Writing ' + sub + ' to row: ' + str(current_row) + ' and column: ' + str(current_column))
            wb = openpyxl.load_workbook(filename)
            sheet = wb.get_sheet_by_name(sheetname)
            sheet.cell(row = current_row, column = current_column).value = sub
            current_column += 1
            wb.save(filename)
            
        elif type(sub) == str:
            print('Writing ' + sub + ' to row: ' + str(current_row) + ' and column: ' + str(current_column))
            wb = openpyxl.load_workbook(filename)
            sheet = wb.get_sheet_by_name(sheetname)
            sheet.cell(row = current_row, column = current_column).value = sub
            current_column += 1
            wb.save(filename)
        else:
            [current_row, current_column] = write_assy_to_excel(sub, current_row, current_column, directory, filename, sheetname)
    


    return [current_row, start_column]
    
    
    
def path_analysis(path, df_port, df_part, df_connector):
    from support_functions import element_idx
    from support_functions import port_between
    ############################################################################
    # check to ensure path is of sufficient length
    if len(path) <= 2:
        return ['Error.', 'Insufficient path information', ['Error'], [], 'Error.', [], False]
    ############################################################################
    # find the first and last part by element ID; check to ensure they are parts        
    first_part = path[0]
    if first_part not in df_part['Elm ID'].values:
        return ['Error.', 'First element not a part.', ['Error'], [], 'Error.', [], False]
    for part in reversed(path):
        if part[0] == '_':
            last_part = part
            break
    if last_part not in df_part['Elm ID'].values:
        return ['Error.', 'Last element not a part.', ['Error'], [], 'Error.', [], False]

    ############################################################################
    # find the set of different port_types used in the path and the set of directions
    # note that if the set of directions is not all one way or a mix of one way /        
    port_types = set()
    path_direction = []
    
    for idx, elem in enumerate(path):
        [elem_idx, elem_type] = element_idx(elem, df_port, df_part, df_connector)
        if elem_type == 'port':
            port_types.add(df_port['Type'][elem_idx])
            path_direction.append(df_port['Direction'][elem_idx])
        elif elem_type in ['connector', 'part'] and idx < (len(path) - 1):
            [next_elem_idx, next_elem_type] = element_idx(path[idx + 1], df_port, df_part, df_connector)
            if next_elem_type in ['connector', 'part']:
                pb = port_between(elem, path[idx + 1], df_port, df_part, df_connector)
                if pb:
                    port_types.add(df_port['Type'][pb[1]])
                    path_direction.append(str(df_port['Direction'][pb[1]]))
                else:
                    port_types.add('No port IDed')
                    path_direction.append('No port IDed')
            else: pass
        else: pass
            
    if len(port_types) > 1 and 'No port IDed' in port_types:
        port_types.remove('No port IDed')

    
    port_direction_conflicts = False
    
    for i, d in enumerate(path_direction):
        if i < (len(path_direction) - 1) and ([d, path_direction[i+1]] in [['in', 'in'], ['out', 'out']]):
            port_direction_conflicts = True
    
    
#    idx = 0
#    while idx <= (len(path)-2): # only need to look at every pair between each setp in the path.  don't need to look at end criteria
#        #find the port between the two components
#        if path[idx + 1] == 'Indeterminate part property from last connector.':
#            pb = port_between(path[idx], path[idx+2], df_port, df_part, df_connector) #returns [port_elm_id, port_idx]; if no port, returns []
#            idx = idx + 2
#        else:
#            pb = port_between(path[idx], path[idx+1], df_port, df_part, df_connector) #returns [port_elm_id, port_idx]; if no port, returns []
#            idx = idx + 1
#                        
#        if pb:# if there is a port between, then add its type to the port types set
#            port_types.add(df_port['Type'][pb[1]])
#            port_direction.add(str(df_port['Direction'][pb[1]]))
#        if not port_types: #if there is no port between, then add 'Indeterminate to the list.
#            port_types.add('No Port - Indeterminate')
#            port_direction.add('No Port - Indeterminate')
#    if len(port_types) > 1 and 'No Port - Indeterminate' in port_types:
#        port_types.remove('No Port - Indterminate')
        
    # check to refine the path of an indeterminate path.
    First_Part_Refiner = []
    if 'Indeterminate part property from last connector.' in path:
        indexes = [i for i in range(len(path)) if path[i] == 'Indeterminate part property from last connector.']
        for i in indexes:
            First_Part_Refiner.append(path[i+1])
    
    English = str(df_part['Name'][element_idx(first_part, df_port, df_part, df_connector)[0]])
    #for fpr in First_Part_Refiner:
    #    English = English + ' of ' + str(df_part['Name'][element_idx(fpr, df_port, df_part, df_connector)[0]])
    English = English + ' connects to ' + str(df_part['Name'][element_idx(last_part, df_port, df_part, df_connector)[0]]) + ' via: '
    for pt in list(port_types):
        English = English + str(pt) + ', '
    English = English + 'port type(s).'
    English.replace(', port type(s)', ' port type(s).')
        

    return [first_part, last_part, list(port_types), First_Part_Refiner, English, path_direction, port_direction_conflicts]
 
 
 
 #this functions takes a connector and a part the connector attaches to and returns the jack according to our modeling pattern
#one flaw of this program is that it assumes you've gotten to a part property who's next layer will be a jack... and that jacks are only modeled as single elements or a jack with a pin as a part property or with a proxy port on the end of the jack
#therefore, it makes sense to only search from a connector and port that you know are at the end of a search... e.g., the part is a CI.  though its possible you could get some weird results.
def physical_jack(connector, part, df_connector, df_part, df_port):
    from support_functions import element_idx
    from support_functions import port_between
    from support_functions import all_connectors_attached_to_port
    
    if connector not in df_connector['Element ID'].values or part not in df_part['Elm ID'].values:
        return {'Error: Connector or Part not in DataFrame.'}
        
    pb = port_between(connector, part, df_port, df_part, df_connector)
    if not pb:
        e_idx = element_idx(part, df_port, df_part, df_connector)[0]
        return {'Direct connection into: ' + str(df_part['Type'][e_idx])}#return the type of the direct connection into the part
        
    port = pb[0]
    port_idx = pb[1]

    # find connectors directly from port 
    s = df_connector['ElmID of Role of Connector Ends (Port)_split']
    idx = list(s[s.apply(lambda i: df_port['Elm ID'][port_idx] in i)].index.values)
    direct_connectors_from_pb = []
    for i in idx:
        if df_connector['Owner'][i] == df_port['Owner'][port_idx]: direct_connectors_from_pb.append(i)
    
    # find the index and Elm ID of each "subport"
    s = df_port['Owner']
    idx = list(s[s.apply(lambda i: df_port['Type'][port_idx] == i)].index.values)
    subports = []
    for i in idx: subports.append(df_port['Elm ID'][i])

    # find each connector attached to the subport that is owned by the primary port    
    rel_conn = []
    for sp in subports:
        sp_conn = all_connectors_attached_to_port(sp, df_connector)
        for c in sp_conn[1]:
            if df_connector['Owner'][c] == df_port['Owner'][port_idx]: rel_conn.append(c)
    
    
    #find inherited subports
    inherited_subports = df_port['Inherited Subport Elm ID_split'][port_idx]
    subport_conn = []
    for isp in inherited_subports:
        isp_conn = all_connectors_attached_to_port(isp, df_connector)
        for c in isp_conn[1]:
            if port in df_connector['Part with Port for Connector Ends ElmID_split'][c]: subport_conn.append(c)
    #print('Subport_conn: ')
    #print(subport_conn)
                
    
    #for each remaining connector build the set of roles to which it is attached:
    # three ways: attached directly to an element
    # attached to port attached to element
    
    roles_to_check = set()  
    for rc in rel_conn:
        rc_idx = element_idx(df_connector['Element ID'][rc], df_port, df_part, df_connector)[0]
        rc_roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][rc_idx]
        for r in rc_roles:
            if r not in subports: roles_to_check.add(r)
    
    for spc in subport_conn:
        spc_roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][spc]
        for s in spc_roles:
            if s not in inherited_subports: roles_to_check.add(s)
    
    for dc in direct_connectors_from_pb:
        dc_roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][dc]
        for d in dc_roles:
            if d not in pb: roles_to_check.add(d)
    
    jacks = set()
    #jackNames = set()
    for r in roles_to_check:
        [r_idx, r_type] = element_idx(r, df_port, df_part, df_connector)
        #case 1: role is a part
        if r_type == 'part':
            if df_part['Owner'][r_idx] == df_port['Owner'][port_idx]:
                jacks.add(df_part['Type'][r_idx])
                #jackNames.add(str(df_part['Name'][r_idx]) + ':' + str(df_part['Type'][r_idx]))
            else:
                #find part that owns r & is owned by df_port['Owner'][port_idx]
                s = df_part['Type']
                idx = list(s[s.apply(lambda i: df_part['Owner'][r_idx] == i)].index.values)
                sup_part_idx = []
                for i in idx:
                    if df_part['Owner'][i] == df_port['Owner'][port_idx]: sup_part_idx.append(i)
                for i in sup_part_idx: 
                    jacks.add(df_part['Type'][i])
                    #jackNames.add(str(df_part['Name'][i]) + ':' + str(df_part['Type'][i]))

        elif r_type == 'port':
            jacks.add(df_port['Owner'][r_idx])
    
    return(jacks)
        
        
def path_context(path, df_connector, df_part, df_port):
    
    from support_functions import element_idx
    from support_functions import part_between
    
    if element_idx(path[0], df_port, df_part, df_connector)[1] != 'part': return [[], [], [], [], '']
    
    A_context_ElmID = []
    A_context_idx = []
    B_context_ElmID = []
    B_context_idx = []
    
    A_or_B = 'A'
    
    for idx, element in enumerate(path):
        print('Path context: ' + str(idx) + ': ' + element)
        
        [e_idx, e_type] = element_idx(element, df_port, df_part, df_connector)
        
        if e_type == 'part':
            if A_or_B == 'A':
                A_context_ElmID.append(element)
                A_context_idx.append(e_idx)

            else:
                B_context_ElmID.append(element)
                B_context_idx.append(e_idx)

                
        elif e_type == 'connector':
            #look at the part the connector is in
            
            role_owners = []
            for e in df_connector['ElmID of Role of Connector Ends (Port)_split'][e_idx]:
                [role_idx, role_type] = element_idx(e, df_port, df_part, df_connector)
                if role_type == 'port': role_owners.append(df_port['Owner'][role_idx])
                elif role_type == 'part': role_owners.append(df_part['Type'][role_idx])

                
            if df_connector['Owner'][e_idx] not in role_owners: 

                A_or_B = 'B'
                top_type = df_connector['Owner'][e_idx]
                top_elmID = df_connector['Element ID'][e_idx]


            next_element = path[idx + 1]
            [next_e_idx, next_e_type] = element_idx(next_element, df_port, df_part, df_connector)

                        
            if next_e_type == 'connector':
                pb = part_between(element, next_element, df_connector, df_part, df_port)
                if len(pb) == 1:
                    [pb_idx, pb_type] = element_idx(pb[0], df_port, df_part, df_connector)
                    if A_or_B == 'A':
                        A_context_ElmID.append(pb[0])
                        A_context_idx.append(pb_idx)
                    elif A_or_B == 'B':
                        B_context_ElmID.append(pb[0])
                        B_context_idx.append(pb_idx)
                
            
        elif e_type == 'port':
            print('Port in path at index: ' + str(e_idx))
            
        else:
            pass
            
    return [A_context_ElmID, A_context_idx, B_context_ElmID, B_context_idx, top_type, top_elmID]



# reads an original CI Table and 1 or more changed CI tables with inputs as indicated below:
    #directory = string of directory where files are held, e.g. '/Users/gillespie/Desktop'
    #original_file = string with name of original_file, e.g. 'original.xlsx'
    #change_files_list = list with strings of names of change files, e.g. ['change_file1.xlsx', 'change_file2.xlsx']
    #change_authors_list = list with strings of names of authors of change file, e.g. ['Gillespie', 'Cole']
    #date = string with date of change, e.g. '1MAR2018'
    #tabletype = 'Connections' or 'CI' this helps you find the appropriate column 

# returns 4 x variables: [JSON_list, no_auto_change, set_of_conflicting_changes, change_list]
# JSON_list is a list of dictionaries in this format:
     #{'id': magicdrawID as string, 'ops': [{'op': 'replace', 'path': magicdraw path within that element ID, 'value': change value}]} may be turned into a JSON and used in Bjorn's importer macro

# no_auto_change which is a list of change numbers that correspond to the index of the item in the change list (explained below)
# these are changes that were not turned into a JSON dictionary for a variety of reasons (e.g., do not want to try to change a magic draw ID

# set_of_conflicting_changes is a list of lists where each list contains a set of change numbers where multiple authors attempted to try to change the same information.  For example, Gillespie tried to change the Name of the 1 DOF Load Cell Signal Board to cool and Cole tried to change it to not_cool - this will generate a set with indexes that correspond to those two changes

# change list is a list of dictionaries where each dictionary contains: {Change#, Original, Change, Original File, Change File, Change Author, etc...}


def excel_file_difference_JSON(directory, original_file, change_files_list, change_authors_list, date, tabletype, sheet):

    
    import os
    import numpy as np
    import simplejson
    import pandas as pd
    
    if tabletype == 'CI': 
        sr = 1
        uniqueID = 'Element ID'
    elif tabletype == 'Connections': 
        sr = 4
        uniqueID = '' # TBD
    else:
        print('Unknown table type.')
        return 
    
    os.chdir(directory)
    original_df = pd.read_excel(original_file, skiprows = sr, sheetname = sheet)
    change_list = []
    change_num = 1
    # Step 1: Find Changes
    # Input: all workbooks
    # Output DataFrame with Changes
    # Sheet Name, Row#, Cell#, Original_Header, Change_Header, Original Contents, Change Contents, Change File, Change Author, Make Change, Reason

    # iterate
    for CFL_idx in np.arange(len(change_files_list)):
        change_df = pd.read_excel(change_files_list[CFL_idx], skiprows = sr, sheetname = sheet)
        change_author = change_authors_list[CFL_idx]
        
        #iterate through each unique identifier
        for ID in original_df[uniqueID].values:
            print(ID)
            #Find the index for the uniquely identified row in both the original and change data frames
            original_idx = original_df[original_df[uniqueID] == ID].index.tolist()
            change_idx = change_df[change_df[uniqueID] == ID].index.tolist()
            
            if len(original_idx) == 0:
                print('Indexing error.')
                continue
            if len(change_idx) == 0:
                change_list.append({'Change#': change_num, 'ChangeType': 'Non Unique Identifiers', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': ID, 'Change': 'ID not found in change index'})
                change_num += 1
                continue
                
            #error check.  sometimes pandas imports rows in excel with nothing in them.  the unique ID will be NaN; simply skip this row
            if type(original_df[uniqueID][original_idx[0]]) == float and np.isnan(original_df[uniqueID][original_idx[0]]): continue
            # error check. unique identifiers that aren't NaN should only be found once in the unique ID list
            if len(original_idx) > 1:
                change_list.append({'Change#': change_num, 'ChangeType': 'Non Unique Identifiers', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': ID, 'Change': ID})
                change_num += 1
                continue
            
            #find the index for the corresponding file in the change file
            if len(change_idx) != 1:
                change_list.append({'Change#': change_num, 'ChangeType': 'Unique ID Added or Deleted.', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': ID, 'Change': 'Listed ' + str(len(change_idx)) + ' times.'})
                change_num += 1
                continue
            
            #now have unique index for both original and change rows.  compare each row / column pair
            
            for col in original_df.columns:
                print(col)
                if col not in change_df.columns:
                    change_list.append({'Change#': change_num, 'ChangeType': 'Columnn Deleted.', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': col, 'Change': col + ' deleted.'})
                    change_num += 1
                    continue
                
                if original_df[col][original_idx[0]] != change_df[col][change_idx[0]]:
                    if (type(original_df[col][original_idx[0]]) in [float, np.float16, np.float32, np.float64] and np.isnan(original_df[col][original_idx[0]])) and (type(change_df[col][change_idx[0]]) in [float, np.float16, np.float32, np.float64] and np.isnan(change_df[col][change_idx[0]])):
                        continue
                    change_list.append({'Change#': change_num, 'ChangeType': 'Cell Change', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': original_df[col][original_idx[0]], 'Change': change_df[col][change_idx[0]], 'ChangeCol': col, 'ChangeRow': ID})
                    change_num += 1
                    
            # if there are additional columns in the change file
            if not not set(change_df.columns).difference(set(original_df.columns)):
                change_list.append({'Change#': change_num, 'ChangeType': 'Column Addition', 'ChangeAuthor': change_author, 'ChangeFile': change_files_list[CFL_idx], 'ChangeSheet': sheet, 'Original': str(original_df.columns), 'Change': str(set(change_df.columns).difference(set(original_df.columns)))})
                change_num += 1
                
        #return change_list
    
    
    # Step 2: Scrub Changes
    # check to ensure original and change header are same
    # check to see if only one change was made to that cell
    # update data frame to Make Change as boolean
    set_of_conflicting_changes = [] # set of lists where each list is the set of change numbers that 
    multi_change_set = set()
    
    for idx1 in np.arange(len(change_list)):
        if change_list[idx1]['ChangeType'] != 'Cell Change': continue
        multi_change = [change_list[idx1]['Change#']]
        if change_list[idx1]['Change#'] in multi_change_set: continue
        
        for idx2 in np.arange(idx1 + 1, len(change_list)):
            if change_list[idx2]['ChangeType'] != 'Cell Change': continue
            if change_list[idx1]['ChangeCol'] == change_list[idx2]['ChangeCol'] and change_list[idx1]['ChangeRow'] == change_list[idx2]['ChangeRow']:
                multi_change.append(change_list[idx2]['Change#']) 
        
        if len(multi_change) > 1: 
            set_of_conflicting_changes.append(multi_change)
            multi_change_set.union(set(multi_change))
        
    
    
    # Step 3: Build dictionaries for JSON
    #dictionary looks like: {'id': element ID, 'ops': [{"op": 'replace', 'path': 'string', 'value': newvalue}]}
    # Step 4: Output excel from Pandas change log and JSON
    
    JSON_list = []
    no_auto_change = []
    for change in change_list:
        #building JSON only based on 
        if change['ChangeType'] != 'Cell Change' or change['Change#'] in multi_change_set: continue
        
        
        # build JSON based on file type
        if tabletype == 'CI':
            # item
            col_to_path_CI = {'Name': '/name', 'Design Status': '/designStatus', 'Documentation': '/Documentation/Hyperlinks', 
            'Design Review Notes': '/designReviewNotes', 'Open Questions': '/openQuestions', 
            'WBS Number': '/WBS Number', 'Vendor': '/Vendor', 'CI Indicator': '/CI Indicator',
            'Functional Area': '/Functional Area', 'Lead': '/Lead', 'Notes': '/Notes',
            'Current Maturity': '/currentMaturity', 'CI Acronym': '/CI Acronym',
            'Tier': '/Tier'}
            
            col_to_path_Weight = {'Lowest Weight': '/lowerValue', 'Most Likely Weight': '/defaultValue', 'Highest Weight': 'upperValue'}
            
            if change['ChangeCol'] in col_to_path_CI.keys():
                JSON_list.append({"id": str(change['ChangeRow']), "ops": [{"op": 'replace', "path": col_to_path_CI[change['ChangeCol']], "value": change['Change'] }]})
            elif change['ChangeCol'] in col_to_path_Weight.keys():
                original_idx = original_df[original_df[uniqueID] == change['ChangeRow']].index.tolist()
                JSON_list.append({"id": str(original_df['ML Weight ID'][original_idx[0]]), "ops": [{"op": 'replace', "path": col_to_path_Weight[change['ChangeCol']], "value": change['Change'] }]})
            else:
                no_auto_change.append(change['Change#'])
    
    
    return [{"modification targets": JSON_list}, no_auto_change, set_of_conflicting_changes, change_list]

                 
        
    
    
    
    
    # 'Notes' # id: CI element ID, op: 'replace', path: 'System Context/WBS Context/Notes' value: changevalue
    # 'Current Maturity' # id: CI element ID, op: 'replace', path: 'Maturity/Current Maturity' value: changevalue                        --> note as of now, it can only be.
    # 'CI Acronym' # id: CI element ID, op: 'replace', path: 'System Context/Configuration Item/CI Acronym' value: changevalue
    
    
    # 'Lowest Weight' # id: ML Weight ID, op: 'replace', path: 'Lower Value' value: changevalue
    # 'Most Likely Weight' # id: ML Weight ID, op: 'replace', path: 'Default Value' value: changevalue
    # 'Highest Weight' # id: ML Weight ID, op: 'replace', path 'Upper Value' value: changevalue
    



    # 'Kind' # <-- this is based off of base classifier This may be very hard to change.....                                             need to talk to Bjorn
    # 'Basis' ## id: CI element ID, op: 'replace', path: 'Maturity/Basis' value: changevalue                                             --> do we want this to serve as a changelog?  this may start being a basis for some of the work

    # 'Element ID' #No Change
    # 'ML Weight ID' #No Change
    # 'ID return' #No change - IDs of part properties for usages of current
    # 'Assembly Tier'   # these are input from script?  do we want to use this to input them into the model?
    # 'Assembly Tier 2'
    # 'Assembly Tier 3'
    # 'Assembly Tier 4'
    # 'Assembly Tier 5'
    # 'Total Weight' # no change
    
    
    # '#' # No Change
    # 'Quantity in Assembly'# No change - export for manual fix (auto count of usages of current and specializing)
    # 'Usages of current and specializing' # see above
        
    # 'Name' # id: CI element ID, op: 'replace', path: 'Name' value: changevalue
    # 'Design Status' ## id: CI element ID, op: 'replace', path: 'System Context/Design Maturity/Design Status' value: changevalue
    # 'Documentation' # id: CI element ID, op: 'replace', path: 'Documentation/Hyperlinks' value: changevalue
    # 'Design Review Notes'# id: CI element ID, op: 'replace', path: 'System Context/Design Maturity/Design Review Notes' value: changevalue
    # 'Open Questions' # id: CI element ID, op: 'replace', path: 'System Context/Design Maturity/Open Questions' value: changevalue
    # 'WBS Number' # id: CI element ID, op: 'replace', path: 'System Context/WBS Context/WBS Number' value: changevalue
    # 'Vendor' # id: CI element ID, op: 'replace', path: 'System Context/WBS Context/Vendor' value: changevalue
    # 'CI Indicator' # id: CI element ID, op: 'replace', path: 'System Context/Configuration Item/CI Indicator' value: changevalue       --> note as of now, it can only be CI or ASM
    # 'Functional Area' # # id: CI element ID, op: 'replace', path: 'System Context/WBS Context/Functional Area' value: changevalue      --> note as of now, these are limited possibiliteis
    # 'Lead' # id: CI element ID, op: 'replace', path: 'System Context/WBS Context/Lead' value: changevalue
    
    
    
def write_changes(JSON_list, no_auto_change, set_of_conflicting_changes, change_list, directory, excel_filename, json_filename):
    
    import openpyxl
    import numpy as np
    import os
    import json
    
    wb = openpyxl.Workbook()
    os.chdir(directory)

    # write Change List to Excel
    sheet = wb.active
    sheet.title = 'Change List'
    columns = [0]
    curr_row = 2
    for change in change_list:
        keys = change.keys()
        for k in keys:
            if k not in columns:
                columns.append(k)
            k_col = columns.index(k)
            sheet.cell(row = curr_row, column = k_col).value = change[k]
        curr_row += 1
    for i in np.arange(1, len(columns)):
        sheet.cell(row = 1, column = i).value = columns[i]
    

    # write no_auto_change to excel
    nac_sheet = wb.create_sheet(title = 'No Automatic Change Sheet')
    columns = [0]
    curr_row = 2    
    for n in no_auto_change:
        change = change_list[n-1]
        keys = change.keys()
        for k in keys:
            if k not in columns:
                columns.append(k)
            k_col = columns.index(k)
            nac_sheet.cell(row = curr_row, column = k_col).value = change[k]
        curr_row += 1
    for i in np.arange(1, len(columns)):
        nac_sheet.cell(row = 1, column = i).value = columns[i]
    
    # write conflicts to excel
    conflicts_sheet = wb.create_sheet(title = 'Change Conflicts Sheet')
    columns = [0, 'Conflict#']
    curr_row = 2
    conflictNum = 1
    
    for c in set_of_conflicting_changes:    
        conflicts_sheet.cell(row = curr_row, column = 1).value = conflictNum
        for n in c:
            change = change_list[n-1]
            keys = change.keys()
            for k in keys:
                if k not in columns:
                    columns.append(k)
                k_col = columns.index(k)
                conflicts_sheet.cell(row = curr_row, column = k_col).value = change[k]
            curr_row += 1
        curr_row += 1
        conflictNum += 1
    for i in np.arange(1, len(columns)):
        conflicts_sheet.cell(row = 1, column = i).value = columns[i]
        
        
    wb.save(excel_filename)
    
    with open(json_filename, 'w') as outfile:
        json.dump(JSON_list, outfile)
    
    
    #[JSON_list, no_auto_change, set_of_conflicting_changes, change_list]
    
    
    
def PathInEnglish(path, df_part, df_port, df_connector):
    Final_Result = []
    if path == []:
        return ['No path']
        
    for p in path:
        if p[0] != '_':
            Final_Result.append('End of path')
            Final_Result.append(p)
        else:
            if p in df_connector['Element ID'].values:
                e_idx = df_connector[df_connector['Element ID'] == p].index.tolist()
                Final_Result.append('CONNECTOR:')
                Final_Result.append('Element ID: ')
                Final_Result.append(p)
                Final_Result.append('Name: ')
                Final_Result.append(df_connector['Name'][e_idx[0]])
                Final_Result.append('Owner: ')
                Final_Result.append(df_connector['Owner'][e_idx[0]])
                
            elif p in df_part['Elm ID'].values:
                e_idx = df_part[df_part['Elm ID'] == p].index.tolist()
                Final_Result.append('PART: ')
                Final_Result.append('Element ID: ')
                Final_Result.append(p)
                Final_Result.append('Name: ')
                Final_Result.append(df_part['Name'][e_idx[0]])
                Final_Result.append('Type: ')
                Final_Result.append(df_part['Type'][e_idx[0]])
                Final_Result.append('Owner: ')
                Final_Result.append(df_part['Owner'][e_idx[0]])
                Final_Result.append('CI?: ')
                Final_Result.append(df_part['CI Indicator (System Context)'][e_idx[0]])
                
            elif p in df_port['Elm ID'].values:
                e_idx = df_port[df_port['Elm ID'] == p].index.tolist()
                Final_Result.append('PORT')
                Final_Result.append('Element ID: ')
                Final_Result.append(p)
                Final_Result.append('Name: ')
                Final_Result.append(df_connector['Name'][e_idx[0]])
                Final_Result.append('Owner: ')
                Final_Result.append(df_connector['Owner'][e_idx[0]])

            else:
                Final_Result.append('NON PART, PORT, CONNECTOR')
                Final_Result.append(p)

    return Final_Result



def part_between(connector_1, connector_2, df_connector, df_part, df_port):
    from support_functions import element_idx
    
    e_1, type_1 = element_idx(connector_1, df_port, df_part, df_connector)
    e_2, type_2 = element_idx(connector_2, df_port, df_part, df_connector)
    
    if type_1 != 'connector' or type_2 != 'connector':
        return 'Error, elements not connectors'
    
    e1_roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][e_1]
    e2_roles = df_connector['ElmID of Role of Connector Ends (Port)_split'][e_2]
    
    shared_role = set(e1_roles).intersection(set(e2_roles))
    
    if len(shared_role) != 1:
        return 'Error, elements do not share role'
        
    sharedRole_idx, x = element_idx(list(shared_role)[0], df_port, df_part, df_connector)
    
    if df_connector['Owner'][e_1] == df_port['Owner'][sharedRole_idx]:
        innerconnector = e_1
        outerconnector = e_2
    else:
        innerconnector = e_2
        outerconnector = e_1
    
    potential_part = []
    
    for part in df_connector['Part with Port for Connector Ends ElmID_split'][outerconnector]:
        if df_part['Type'][element_idx(part, df_port, df_part, df_connector)[0]] == df_connector['Owner'][innerconnector]: potential_part.append(part)
    
    return potential_part
        
            
def ownCI(element, CI_Set, df_connector, df_part, df_port):
    #Note that this only looks one level deep.  We could build it recursively, 
    #but at this time it doesn't seem necessary
    
    from support_functions import element_idx
    e_idx, e_type = element_idx(element, df_port, df_part, df_connector)
    
    # a non-part can't own another part, return false
    if e_type != 'part': return False 

    # find all parts that have owner whose type is the type of the part
    owned_parts = df_part[df_part['Owner'] == df_part['Type'][e_idx]].index.tolist()
    
    # if there are no owned parts, none are CIs
    if not owned_parts: return False
    
    # check each owned part to see if it is a CI
    for op in owned_parts:
        if df_part['CI Indicator (System Context)'][op] in CI_Set: return True
    
    return False
    