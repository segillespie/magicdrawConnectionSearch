################################################################################
#                   MODIFY THIS AS DESIRED / APPROPRIATE
# model version for the inputs
model_version = 1650
# the directory where all of your query files (search, support_functions, etc.. are stored
query_functions_directory = "c:/Users/stephen.gillespie/Desktop/Mac/"
# the directory where your Magic Draw data exports are stored
input_file_directory = "c:/Users/stephen.gillespie/Desktop/Mac/"
# the directory where you want to save the results of this script
output_file_directory = "c:/Users/stephen.gillespie/Desktop/Mac/" #/MagicDraw_Query/Updated_Model_Output'
# Analyst Name and contactinfo
analyst_name = 'Stephen Gillespie'
analyst_email = 'stephen.gillespie@usma.edu'

PartsFile = 'Model_Part_Properties_(CI Usage)_' + str(model_version) + '.xlsx'
sheet_name = 'Data_Only'

# These are modifiable categories to tag various paths as desired
CI_Set = ['CI', 'Ci', 'ci', 'cI', 'Y', 'y', 'Yes', 'yes', 'YES', 'C', 'c', 'CI - Structural', 'Sub-CI', 'ASM']

# column order for printing to excel
# naming notation is VariableName_col, where VariableName is the equivalent variable that will be written to that column
A_Name_col = 1
A_Type_col = 2
A_FA_col = 3
A_CI_col = 4
A_Kind_col = 5
A_Vendor_col = 6
A_Mult_Lower_col = 7
A_Mult_Upper_col = 8
A_Tier_col = 9
A_Tier2_col = 10
A_Tier3_col = 11
A_Tier4_col = 12
A_Tier5_col = 13
A_Tier1_col = 14
item_col = 15
A_ElmID_col = 16
MultiUseIndicator_col = 17
#make this the equal to the highest of the above; assists with formatting
highest_col_num = 17

#A_ML_Weight_col = 7


################################################################################
# first import import necessary packages etc..
import os
os.chdir(query_functions_directory)
import numpy as np
import openpyxl
import datetime
from support_functions import XLSX_to_DF
from support_functions import getAssyTier

################################################################################
# second convert your inputs into pandas dataframes for analysis
print('Reading input excels.')
df_part = XLSX_to_DF(input_file_directory, PartsFile, sheet_name)

################################################################################
#Set up the Excel Workbook For Desired Output
wb = openpyxl.Workbook()
sheet = wb.active
current_row = 1 
current_column = 1
item = 0
bold14 = openpyxl.styles.Font(size=14, bold=True)
ital11 = openpyxl.styles.Font(italic=True)
sheet.title = 'CI Usage Table'

#Top Level Headers - Add top level headers / info
print('Writing row: ' + str(current_row))
sheet.cell(row = current_row, column = 1).value = 'CI Usage Table for TALOS.\nGenerated from MagicDraw model version: ' + str(model_version) + datetime.datetime.now().strftime(' on %d%b%Y at %H:%M.') + '\nPOC: ' + analyst_name + ' ' + analyst_email + '\nNote: "nan" indicates information not contained in the model.'
sheet.cell(row = current_row, column = 1).font = ital11
sheet.cell(row = current_row, column = 1).alignment = openpyxl.styles.Alignment(wrapText=True)
sheet.row_dimensions[current_row].height = 65  
sheet.merge_cells('A'+str(current_row)+':J'+str(current_row))
current_row += 1

print('Writing row: ' + str(current_row))
sheet.cell(row = current_row, column = item_col).value = 'Data Frame Index #'
sheet.cell(row = current_row, column = A_Name_col).value = 'Usage Name'
sheet.cell(row = current_row, column = A_Type_col).value = 'Usage CI Type'
sheet.cell(row = current_row, column = A_ElmID_col).value = 'Usage MagicDraw ID'
sheet.cell(row = current_row, column = A_FA_col).value = 'Usage Functional Area'
sheet.cell(row = current_row, column = A_Kind_col).value = 'Usage Kind'
sheet.cell(row = current_row, column = A_Tier1_col).value = 'Usage Assembly Tier 1'
sheet.cell(row = current_row, column = A_Tier2_col).value = 'Usage Assembly Tier 2'
sheet.cell(row = current_row, column = A_Tier3_col).value = 'Usage Assembly Tier 3'
sheet.cell(row = current_row, column = A_Tier4_col).value = 'Usage Assembly Tier 4'
sheet.cell(row = current_row, column = A_Tier5_col).value = 'Usage Assembly Tier 5'
sheet.cell(row = current_row, column = A_Tier_col).value = 'Usage Assembly Tier'
sheet.cell(row = current_row, column = A_Vendor_col).value = 'Usage Vendor'
#sheet.cell(row = current_row, column = A_ML_Weight_col).value = 'Usage Weight (lbs)'
sheet.cell(row = current_row, column = A_CI_col).value = 'CI Indicator'
sheet.cell(row = current_row, column = A_Mult_Lower_col).value = 'Usage Lower Multiplicity'
sheet.cell(row = current_row, column = A_Mult_Upper_col).value = 'Usage Upper Multiplicity'
sheet.cell(row = current_row, column = MultiUseIndicator_col).value = 'Multi Use Indicator'


sheet.freeze_panes = 'A' + str(current_row + 1)
for col in np.arange(1, highest_col_num + 1): sheet.cell(row = current_row, column = col).font = bold14

################################################################################
# Write relevant data for each part property that is a CI, Sub-CI, or ASM
for i in np.arange(len(df_part)):
    
    if df_part['CI Indicator (System Context)'][i] in CI_Set and not df_part['Is Abstract'][i]: #and not df_part['Is Software'][i] and not df_part['Is Wire Harness'][i]: 
        print('Reading index: ' + str(i))
        owner_idx = df_part[df_part['Type'] == df_part['Owner'][i]].index.tolist()
        
        if df_part['Owner'][i] == 'FULL TALOS Assembly': owner_idx.append('FULL TALOS Assembly')
            
        for owner in owner_idx:
            current_row = current_row + 1
            ####################################################################
            # Inputs that don't change regardless of number of times part is represented
            # Index in Data Frame
            sheet.cell(row = current_row, column = item_col).value = i 
            #Type
            sheet.cell(row = current_row, column = A_Type_col).value = str(df_part['Type'][i])
            # CI A Functional Area
            #Functional Area
            sheet.cell(row = current_row, column = A_FA_col).value = str(df_part['Functional Area (System Context)'][i])
            #Part Property Element ID
            sheet.cell(row = current_row, column = A_ElmID_col).value = str(df_part['Elm ID'][i])
            #Kind
            sheet.cell(row = current_row, column = A_Kind_col).value = str(df_part['Kind'][i])
            #Vendor
            sheet.cell(row = current_row, column = A_Vendor_col).value = str(df_part['Vendor'][i])
            # Most Likely Weight
            #sheet.cell(row = current_row, column = A_ML_Weight_col).value = str(df_part['Most Likely Weight_split'][i])
            # CI Indicator
            sheet.cell(row = current_row, column = A_CI_col).value = str(df_part['CI Indicator (System Context)'][i])
            #multiplicity
            if df_part['Multiplicity'][i] == '(Unspecified)':
                sheet.cell(row = current_row, column = A_Mult_Lower_col).value = 1
                sheet.cell(row = current_row, column = A_Mult_Upper_col).value = 1
            elif type(df_part['Multiplicity'][i]) == str and '..' in df_part['Multiplicity'][i]:
                HL_Mult = df_part['Multiplicity'][i].split('..')
                sheet.cell(row = current_row, column = A_Mult_Lower_col).value = HL_Mult[0]
                sheet.cell(row = current_row, column = A_Mult_Upper_col).value = HL_Mult[1]
            else:
                sheet.cell(row = current_row, column = A_Mult_Lower_col).value = str(df_part['Multiplicity'][i])
                sheet.cell(row = current_row, column = A_Mult_Upper_col).value = str(df_part['Multiplicity'][i])
        
            ####################################################################
            # Variations depending on number of owner's usage as part
            if len(owner_idx) == 1:
                #Name is the name of the part
                sheet.cell(row = current_row, column = A_Name_col).value = str(df_part['Name'][i]) # CI A - Name
                # Tiers
                tiers = getAssyTier(df_part['Elm ID'][i], df_part)
                tiers_len = len(tiers)
                sheet.cell(row = current_row, column = A_Tier_col).value = tiers_len
                sheet.cell(row = current_row, column = A_Tier1_col).value = tiers[0]
                if tiers_len >= 2: sheet.cell(row = current_row, column = A_Tier2_col).value = tiers[1]
                if tiers_len >= 3: sheet.cell(row = current_row, column = A_Tier3_col).value = tiers[2]
                if tiers_len >= 4: sheet.cell(row = current_row, column = A_Tier4_col).value = tiers[3]
                if tiers_len >= 5: sheet.cell(row = current_row, column = A_Tier5_col).value = tiers[4]
                # Indicator
                sheet.cell(row = current_row, column = MultiUseIndicator_col).value = 'False'
            elif len(owner_idx) > 1 and owner != 'FULL TALOS Assembly':
                #Name of is the name of the part + name of the owning part
                sheet.cell(row = current_row, column = A_Name_col).value = str(df_part['Name'][i]) + ' / ' + str(df_part['Name'][owner])
                #Tiers
                tiers = getAssyTier(df_part['Elm ID'][owner], df_part)
                tiers.append(df_part['Type'][owner])
                tiers_len = len(tiers)
                sheet.cell(row = current_row, column = A_Tier_col).value = tiers_len
                sheet.cell(row = current_row, column = A_Tier1_col).value = tiers[0]
                if tiers_len >= 2: sheet.cell(row = current_row, column = A_Tier2_col).value = tiers[1]
                if tiers_len >= 3: sheet.cell(row = current_row, column = A_Tier3_col).value = tiers[2]
                if tiers_len >= 4: sheet.cell(row = current_row, column = A_Tier4_col).value = tiers[3]
                if tiers_len >= 5: sheet.cell(row = current_row, column = A_Tier5_col).value = tiers[4]               
                # Indicator
                sheet.cell(row = current_row, column = MultiUseIndicator_col).value = 'True'    


################################################################################
# Adjust column widths so that it is readable
columnletters = ('', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ')
sheet.column_dimensions[columnletters[A_Type_col]].width = 40
sheet.column_dimensions[columnletters[A_Name_col]].width = 35.43
sheet.column_dimensions[columnletters[A_FA_col]].width = 14
sheet.column_dimensions[columnletters[A_Kind_col]].width = 28.86
sheet.column_dimensions[columnletters[A_Tier_col]].width = 5
sheet.column_dimensions[columnletters[A_Tier2_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier3_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier4_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier5_col]].width = 36.29
sheet.column_dimensions[columnletters[item_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Vendor_col]].width = 19.71
#sheet.column_dimensions[columnletters[A_ML_Weight_col]].width = 19.71
sheet.column_dimensions[columnletters[A_CI_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Tier_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Tier_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Tier1_col]].width = 19.71
sheet.column_dimensions[columnletters[A_ElmID_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Mult_Lower_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Mult_Upper_col]].width = 19.71
sheet.column_dimensions[columnletters[MultiUseIndicator_col]].width = 19.71






os.chdir(output_file_directory)
wb.save('CI_Usage_Table_' + str(model_version) + '.xlsx')

################################################################################
# Assembly Build Option 1

from support_functions import list_or_set_to_string
tiers_df = XLSX_to_DF(output_file_directory, 'CI_Usage_Table_' + str(model_version) + '.xlsx', 'CI Usage Table', 1)
wb.create_sheet(index=2, title='CI Tiers')
tiers_ws = wb.get_sheet_by_name('CI Tiers')
current_row = 1

tiers_ws.cell(row = current_row, column = 1).value = 'CI Name'
tiers_ws.cell(row = current_row, column = 2).value = 'Assembly Tier'
tiers_ws.cell(row = current_row, column = 3).value = 'Assembly Tier 2'
tiers_ws.cell(row = current_row, column = 4).value = 'Assembly Tier 3'
tiers_ws.cell(row = current_row, column = 5).value = 'Assembly Tier 4'
tiers_ws.cell(row = current_row, column = 6).value = 'Assembly Tier 5'

for CI in set(tiers_df['Usage CI Type']):
    current_row += 1
    print(CI)
    tiers_ws.cell(row = current_row, column = 1).value = str(CI)
    
    usage_idx = tiers_df[tiers_df['Usage CI Type'] == CI].index.tolist()
    
    tier2_set = set()
    tier3_set = set()
    tier4_set = set()
    tier5_set = set()
    
    for i in usage_idx: 
        tier2_set.add(tiers_df['Usage Assembly Tier 2'][i])
        tier3_set.add(tiers_df['Usage Assembly Tier 3'][i])
        tier4_set.add(tiers_df['Usage Assembly Tier 4'][i])
        tier5_set.add(tiers_df['Usage Assembly Tier 5'][i])
    
    tiers_ws.cell(row = current_row, column = 3).value = list_or_set_to_string(tier2_set, empty_message = '', delim = ', ')
    tiers_ws.cell(row = current_row, column = 4).value = list_or_set_to_string(tier3_set, empty_message = '', delim = ', ')
    tiers_ws.cell(row = current_row, column = 5).value = list_or_set_to_string(tier4_set, empty_message = '', delim = ', ')
    tiers_ws.cell(row = current_row, column = 6).value = list_or_set_to_string(tier5_set, empty_message = '', delim = ', ')
    
    if np.nan in tier2_set: tiers_ws.cell(row = current_row, column = 2).value = '1'
    elif np.nan in tier3_set: tiers_ws.cell(row = current_row, column = 2).value = '2'
    elif np.nan in tier4_set: tiers_ws.cell(row = current_row, column = 2).value = '3'
    elif np.nan in tier5_set: tiers_ws.cell(row = current_row, column = 2).value = '4'
    else: tiers_ws.cell(row = current_row, column = 2).value = '5'
    

    
wb.save('CI_Usage_Table_' + str(model_version) + '.xlsx')


################################################################################
# Assembly Build Option 2

from support_functions import list_or_set_to_string
import numpy as np

wb.create_sheet(index=2, title='CI Tiers-2')
tiers_ws = wb.get_sheet_by_name('CI Tiers-2')
current_row = 1

tiers_ws.cell(row = current_row, column = 1).value = 'CI Name'
tiers_ws.cell(row = current_row, column = 2).value = 'Assembly Tier'
tiers_ws.cell(row = current_row, column = 3).value = 'Assembly Tier 2'
tiers_ws.cell(row = current_row, column = 4).value = 'Assembly Tier 3'
tiers_ws.cell(row = current_row, column = 5).value = 'Assembly Tier 4'
tiers_ws.cell(row = current_row, column = 6).value = 'Assembly Tier 5'

for CI in set(tiers_df['Usage CI Type']):
    current_row += 1
    print(CI)
    tiers_ws.cell(row = current_row, column = 1).value = str(CI)
    
    usage_idx = tiers_df[tiers_df['Usage CI Type'] == CI].index.tolist()
    
    tier2_set = set()
    tier3_set = set()
    tier4_set = set()
    tier5_set = set()
    
    for i in usage_idx: 
        tier2_set.add(tiers_df['Usage Assembly Tier 2'][i])
        tier3_set.add(tiers_df['Usage Assembly Tier 3'][i])
        tier4_set.add(tiers_df['Usage Assembly Tier 4'][i])
        tier5_set.add(tiers_df['Usage Assembly Tier 5'][i])
    
    if len(tier2_set) == 1: tiers_ws.cell(row = current_row, column = 3).value = list_or_set_to_string(tier2_set, empty_message = '', delim = ', ')
    if len(tier3_set) == 1: tiers_ws.cell(row = current_row, column = 4).value = list_or_set_to_string(tier3_set, empty_message = '', delim = ', ')
    if len(tier4_set) == 1: tiers_ws.cell(row = current_row, column = 5).value = list_or_set_to_string(tier4_set, empty_message = '', delim = ', ')
    if len(tier5_set) == 1: tiers_ws.cell(row = current_row, column = 6).value = list_or_set_to_string(tier5_set, empty_message = '', delim = ', ')
    
    if np.nan in tier2_set: tiers_ws.cell(row = current_row, column = 2).value = '1'
    elif np.nan in tier3_set: tiers_ws.cell(row = current_row, column = 2).value = '2'
    elif np.nan in tier4_set: tiers_ws.cell(row = current_row, column = 2).value = '3'
    elif np.nan in tier5_set: tiers_ws.cell(row = current_row, column = 2).value = '4'
    else: tiers_ws.cell(row = current_row, column = 2).value = '5'

    
wb.save('CI_Usage_Table_' + str(model_version) + '.xlsx')