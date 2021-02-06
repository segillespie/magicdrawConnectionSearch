################################################################################
#                   MODIFY THIS AS DESIRED / APPROPRIATE
# model version for the inputs
model_version = 1650
# the directory where all of your query files (search, support_functions, etc.. are stored
query_functions_directory = "c:/Users/stephen.gillespie/Desktop/Mac/"
# the directory where your Magic Draw data exports are stored
input_file_directory = "c:/Users/stephen.gillespie/Desktop/Mac/"
# the directory where you want to save the results of this script
output_file_directory = "c:/Users/stephen.gillespie/Desktop/Mac/" #/MagicDraw_Query/Updated_Model_Output'
# Analyst Name and contactinfo
analyst_name = 'Stephen Gillespie'
analyst_email = 'stephen.gillespie@usma.edu'

# if true, this hides all references to magicdraw stereotypes / element IDs, etc... - namely rows that for the headers using MagicDraw terms and columns that are filled with elmIDs 
hidemagicdraw = False 


PortFile = 'Model_Ports_' + str(model_version) + '.xlsx'
ConnectorsFile = 'Model_Connectors_' + str(model_version) + '.xlsx'
PartsFile = 'Model_Part_Properties_' + str(model_version) +'.xlsx'
sheet_name = 'Data_Only'

# These are modifiable categories to tag various paths as desired
CI_Set = ['CI', 'Ci', 'ci', 'cI', 'Y', 'y', 'Yes', 'yes', 'YES', 'C', 'c', 'CI - Structural', 'Sub-CI']
physicalIF = ['Physical IF', 'Bolt', 'Velcro']
logicalIF = ['RS 485', 'RS 485 + Safety Enable', 'RS-485', 'Gigabit Ethernet', 'Analog Signal IF', 'TALOS Cooling System Signal IF', 'GigE', 'Analog IF', 'USB IF', 'USB', 'USB 2.0', 'USB 3.0', 'Discrete IF', 'Analog', 'Analog IF', 'eStop IF', 'Ethernet H.264 Video IF', 'RS-485 + Enable Gate', 'TALOS Cooling System Signal IF', '3-Phase Connection IF', 'I2C', 'I2S 2-Way IF', 'MIPI CSI-2', 'Synchronous Serial IF', 'Thermistor IF', 'Transducer IF', 'Vent Tach IF', 'GPS Signal IF', 'Radio Signal IF', 'USB H.264 Video', 'Valve Driver IF', 'Serial Data iF', 'MIPI DSI-2', 'Bluetooth', 'Display Port', 'PCI Express', 'PHYS Data', 'Serial ATA', ]                   
electricalIF = ['60.0 V IF', '48.0 V IF', '16.0 V IF', 'USB IF', 'USB', '5.0 V IF', '3.3 V IF', '1.8 V IF', 'AC IF', 'USB 3.0', '16.0 V Joint Controller', 'Two Pin Power IF']
coolantIF = ['Cooling Interface']

#connector type categories
#CT_Physical = []
#CT_Logical = ['RS-485 Connection', 'RS-485 + Enable Gate Connection',  'PrinterUSBSocketDelegationJunction',  'PCI Express Connection',  'Transducer Connection', 'Gigabit Ethernet Connection', 'SSI Connection',  'Serial Data Connection',  'PrinterUSBSocket2PrinterUSBPlugJunction', 'ComputerUSBInterface2PrinterUSBInterfaceJunction', 'PrinterUSBPlugDelegationJunction', 'Data Connection',  'Thermistor Connection', '<Interface>Junction', 'ComputerUSBSocket2ComputerUSBPlugJunction', 'ComputerUSBSocketDelegationJunction',  'ComputerUSBPlugDelegationJunction', 'RS-485 Connection']
#CT_Electrical = [ 'High Power Connection', 'Low Power Connection',  'AC Connection', 'Low Power 2-pin Connection']
#CT_Other = []


# column order for printing to excel
# naming notation is VariableName_col, where VariableName is the equivalent variable that will be written to that column
A_Type_col = 1
A_Name_col = 2
B_Name_col = 3
A_PortName_col = 4
B_PortName_col = 5
A_PhysicalJack_col = 6
B_PhysicalJack_col = 7
allocatedWH_name_col = 8
B_Type_col = 9
A_Owner_col = 10
B_Owner_col = 11
A_FA_col = 12
B_FA_col = 13
Interface_Category_col = 14
Port_Types_col = 15
A_Kind_col = 16
A_Direction_col = 17
B_Direction_col = 18
A_Tier_col = 19
A_Tier2_col = 20
A_Tier3_col = 21
A_Tier4_col = 22
A_Tier5_col = 23
sentence_col = 24
item_col = 25

A_Tier1_col = 26
A_ElmID_col = 27
B_ElmID_col = 28
path_col = 29
A_PortElmID_col = 30
B_PortElmID_col = 31
path_direction_col = 32
top_level_col = 33
allocatedWH_elmID_col = 34
A_Context_col = 35
B_Context_col = 36
top_elmID_col = 37
B_Tier1_col = 38
B_Tier2_col = 39
A_CI_col = 40
B_CI_col = 41


#make this the equal to the highest of the above; assists with formatting
highest_col_num = 41


################################################################################
# first import import necessary packages etc..
import os
os.chdir(query_functions_directory)
from search import search
from support_functions import path_analysis
import numpy as np
import openpyxl
import datetime
from support_functions import physical_jack
from support_functions import element_idx
from support_functions import list_or_set_to_string
from support_functions import port_between
from support_functions import context_in_english
from support_functions import XLSX_to_DF
from support_functions import path_context
from support_functions import allocatedWireHarness
from support_functions import getAssyTier

################################################################################
# second convert your inputs into pandas dataframes for analysis
print('Reading input excels.')
df_part = XLSX_to_DF(input_file_directory, PartsFile, sheet_name)
df_port = XLSX_to_DF(input_file_directory, PortFile, sheet_name)
df_connector = XLSX_to_DF(input_file_directory, ConnectorsFile, sheet_name)

################################################################################
#Set up the Excel Workbook For Desired Output
wb = openpyxl.Workbook()
sheet = wb.active
current_row = 1 
current_column = 1
item = 0
bold14 = openpyxl.styles.Font(size=14, bold=True)
ital11 = openpyxl.styles.Font(italic=True)
sheet.title = 'Connections'

#Top Level Headers - Add top level headers / info
print('Writing row: ' + str(current_row))
sheet.cell(row = current_row, column = 1).value = 'One Step Connection List CI-to-CI for TALOS.\nGenerated from MagicDraw model version: ' + str(model_version) + datetime.datetime.now().strftime(' on %d%b%Y at %H:%M.') + '\nPOC: ' + analyst_name + ' ' + analyst_email + '\nNote: "nan" indicates information not contained in the model.'
sheet.cell(row = current_row, column = 1).font = ital11
sheet.cell(row = current_row, column = 1).alignment = openpyxl.styles.Alignment(wrapText=True)
sheet.row_dimensions[current_row].height = 65  
sheet.merge_cells('A'+str(current_row)+':J'+str(current_row))
current_row += 1

# Magic Draw Headers
print('Writing row: ' + str(current_row))
sheet.cell(row =current_row, column = 1).value = 'MagicDraw Model Headers.'
if hidemagicdraw: sheet.row_dimensions[current_row].hidden = True
for col in np.arange(1, highest_col_num + 1): sheet.cell(row = current_row, column = col).font = bold14
current_row += 1

print('Writing row: ' + str(current_row))
sheet.cell(row = current_row, column = item_col).value = 'Item Number'
sheet.cell(row = current_row, column = A_Name_col).value = 'CI A - Name'
sheet.cell(row = current_row, column = A_Type_col).value = 'CI A - Type'
sheet.cell(row = current_row, column = A_Owner_col).value = 'CI A - Owner'
sheet.cell(row = current_row, column = A_Context_col).value = 'CI A - Context'
sheet.cell(row = current_row, column = A_PhysicalJack_col).value = 'CI A - Physical Jack'
sheet.cell(row = current_row, column = A_PortName_col).value = 'CI A - Port Name'
sheet.cell(row = current_row, column = A_Direction_col).value = 'CI A - Direction'
sheet.cell(row = current_row, column = Port_Types_col).value = 'Port Type(s)'
sheet.cell(row = current_row, column = B_Name_col).value = 'CI B - Name'
sheet.cell(row = current_row, column = B_Type_col).value = 'CI B - Type'
sheet.cell(row = current_row, column = B_Owner_col).value = 'CI B - Owner'
sheet.cell(row = current_row, column = B_Context_col).value = 'CI B - Context'
sheet.cell(row = current_row, column = B_PhysicalJack_col).value = 'CI B - Physical Jack'
sheet.cell(row = current_row, column = B_PortName_col).value = 'CI B - Port Name'
sheet.cell(row = current_row, column = B_Direction_col).value = 'CI B - Direction'
sheet.cell(row = current_row, column = sentence_col).value = 'Sentence Description'
sheet.cell(row = current_row, column = A_ElmID_col).value = 'CI A - Element ID'
sheet.cell(row = current_row, column = B_ElmID_col).value = 'CI B - Element ID'
sheet.cell(row = current_row, column = path_col).value = 'Path'
sheet.cell(row = current_row, column = A_PortElmID_col).value = 'Port A - Element ID'
sheet.cell(row = current_row, column = B_PortElmID_col).value = 'Port B - Element ID'
sheet.cell(row = current_row, column = Interface_Category_col).value = 'Interface Category'
sheet.cell(row = current_row, column = A_FA_col).value = 'CI A - FA'
sheet.cell(row = current_row, column = B_FA_col).value = 'CI B - FA'
sheet.cell(row = current_row, column = path_direction_col).value = 'Path Direction'
sheet.cell(row = current_row, column = top_level_col).value = 'Highest level assembly for path'
sheet.cell(row = current_row, column = top_elmID_col).value = 'Element ID Top Connector'

if hidemagicdraw: sheet.row_dimensions[current_row].hidden = True
for col in np.arange(1, highest_col_num + 1): sheet.cell(row = current_row, column = col).font = bold14
# User Headers
current_row += 1

print('Writing row: ' + str(current_row))
sheet.cell(row =current_row, column = 1).value = 'User Headers.'
if hidemagicdraw: sheet.row_dimensions[current_row].hidden = True
for col in np.arange(1, highest_col_num + 1): sheet.cell(row = current_row, column = col).font = bold14
current_row += 1

print('Writing row: ' + str(current_row))
sheet.cell(row = current_row, column = item_col).value = 'Connection #'
sheet.cell(row = current_row, column = A_Name_col).value = 'Part A Name'
sheet.cell(row = current_row, column = A_Type_col).value = 'Part A CI'
sheet.cell(row = current_row, column = A_Owner_col).value = 'Part A Assembly'
sheet.cell(row = current_row, column = A_Context_col).value = 'Part A Context'
sheet.cell(row = current_row, column = A_PhysicalJack_col).value = 'Part A Physical Jack'
sheet.cell(row = current_row, column = A_PortName_col).value = 'Part A Functional Interface Type'
sheet.cell(row = current_row, column = A_Direction_col).value = 'Part A Jack Flow Direction'
sheet.cell(row = current_row, column = Port_Types_col).value = 'Connection Type(s)'
sheet.cell(row = current_row, column = B_Name_col).value = 'Part B Name'
sheet.cell(row = current_row, column = B_Type_col).value = 'Part B CI'
sheet.cell(row = current_row, column = B_Owner_col).value = 'Part B Assembly'
sheet.cell(row = current_row, column = B_Context_col).value = 'Part B Context'
sheet.cell(row = current_row, column = B_PhysicalJack_col).value = 'Part B Physical Jack'
sheet.cell(row = current_row, column = B_PortName_col).value = 'Part B Functional Interface Type'
sheet.cell(row = current_row, column = B_Direction_col).value = 'Part B Jack Flow Direction'
sheet.cell(row = current_row, column = sentence_col).value = 'Sentence Description'
sheet.cell(row = current_row, column = A_ElmID_col).value = 'Part A MagicDraw ID'
sheet.cell(row = current_row, column = B_ElmID_col).value = 'Part B MagicDraw ID'
sheet.cell(row = current_row, column = path_col).value = 'Magic Draw Path from Part A to Part B' 
sheet.cell(row = current_row, column = A_PortElmID_col).value = 'Magic Draw Port A ID'
sheet.cell(row = current_row, column = B_PortElmID_col).value = 'Magic Draw Port B ID'
sheet.cell(row = current_row, column = Interface_Category_col).value = 'Connection Category'
sheet.cell(row = current_row, column = A_FA_col).value = 'Part A Functional Area'
sheet.cell(row = current_row, column = B_FA_col).value = 'Part B Functional Area'
sheet.cell(row = current_row, column = path_direction_col).value = 'Path Direction'
sheet.cell(row = current_row, column = top_level_col).value = 'Highest level assembly for path'
sheet.cell(row = current_row, column = allocatedWH_elmID_col).value = 'Allocated Wire Harnesses MagicDraw ID'
sheet.cell(row = current_row, column = allocatedWH_name_col).value = 'Allocated Wire Harnesses Names'
sheet.cell(row = current_row, column = A_Kind_col).value = 'Part A Kind'
sheet.cell(row = current_row, column = A_Tier1_col).value = 'Part A Assembly Tier 1'
sheet.cell(row = current_row, column = A_Tier2_col).value = 'Part A Assembly Tier 2'
sheet.cell(row = current_row, column = A_Tier3_col).value = 'Part A Assembly Tier 3'
sheet.cell(row = current_row, column = A_Tier4_col).value = 'Part A Assembly Tier 4'
sheet.cell(row = current_row, column = A_Tier5_col).value = 'Part A Assembly Tier 5'
sheet.cell(row = current_row, column = A_Tier_col).value = 'Part A Assembly Tier'
sheet.cell(row = current_row, column = top_elmID_col).value = 'Element ID Top Connector'
sheet.cell(row = current_row, column = B_Tier1_col).value = 'Part B Assembly Tier 1'
sheet.cell(row = current_row, column = B_Tier2_col).value = 'Part B Assembly Tier 2'
sheet.cell(row = current_row, column = A_CI_col).value = 'Part A CI Indicator'
sheet.cell(row = current_row, column = B_CI_col).value = 'Part B CI Indicator'

sheet.freeze_panes = 'A' + str(current_row + 1)
for col in np.arange(1, highest_col_num + 1): sheet.cell(row = current_row, column = col).font = bold14

################################################################################
# Third search every element in the df_part dataframe that is identified as a CI
unwrittenPaths = []

for i in np.arange(len(df_part)):
    
    if df_part['CI Indicator (System Context)'][i] in CI_Set and not df_part['Is Software'][i] and not df_part['Is Wire Harness'][i]: #and df_part['Functional Area (System Context)'][i] == 'PWR':
        print('Reading index: ' + str(i))
        
        try: 
            paths = search(df_part['Elm ID'][i], [], df_part, df_connector, df_port)
            
            for path in paths:
                if path[len(path) - 1] in ['End of path, last element is CI.', 'End of path, last element is a CI.', 'End of Path: Element not in Ports, Parts, or Connectors DF', 'End of Path: Part has no connectors', 'End of Path: No connectors from current port and part to another. Not a CI', 'End of path, Last part property has no more connectors, not a CI.', 'End of path, last element is part, but not CI.'] and (len(path) > 2):          
                    A_idx = element_idx(path[0], df_port, df_part, df_connector)[0]
                    B_idx = element_idx(path[len(path)-2], df_port, df_part, df_connector)[0]
                    
                    if df_part['Is Software'][B_idx] or df_part['Is Wire Harness'][B_idx]: continue

                    current_row = current_row + 1

                    item = item + 1
                    pa = path_analysis(path, df_port, df_part, df_connector)
                    
                    try:
                        A_context_ElmID, A_context_idx, B_context_ElmID, B_context_idx, top_type, top_elmID = path_context(path, df_connector, df_part, df_port)
                        B_context_ElmID.reverse()
                        B_context_idx.reverse()
                        A_context_list = []
                        B_context_list = []
                        for a in A_context_idx: A_context_list.append(str(df_part['Name'][a]) + ':' + df_part['Type'][a])
                        A_context_list.append(top_type)
                        for b in B_context_idx: B_context_list.append(str(df_part['Name'][b]) + ':' + df_part['Type'][b])
                        B_context_list.append(top_type)
                    except Exception as e:
                        A_context_list = ['Error on path_context()']
                        B_context_list = ['Error on path_context()']
                        top_type = ['Error on path_context()']
                        top_elmID = ['Error on path_context()']
                    
                    sheet.cell(row = current_row, column = A_Kind_col).value = str(df_part['Kind'][A_idx]) 
                    
                    sheet.cell(row = current_row, column = item_col).value = item # Item Number
                    A_Name = str(df_part['Name'][A_idx])
                    sheet.cell(row = current_row, column = A_Name_col).value = A_Name # CI A - Name
                    
                    A_Type = str(df_part['Type'][A_idx])
                    sheet.cell(row = current_row, column = A_Type_col).value = A_Type# CI A - Type
                    
                    A_Owner = str(df_part['Owner'][A_idx])
                    sheet.cell(row = current_row, column = A_Owner_col).value = A_Owner #'CI A - Owner'
                    
                    A_Context = list_or_set_to_string(A_context_list, 'No context', ' , ')
                    sheet.cell(row = current_row, column = A_Context_col).value = A_Context # CI A - Context
                    
                    A_PhysicalJack = list_or_set_to_string(physical_jack(path[1], path[0], df_connector, df_part, df_port), 'No jack identified.')
                    sheet.cell(row = current_row, column = A_PhysicalJack_col).value = A_PhysicalJack#'CI A - Physical Jack'
                    
                    sheet.cell(row = current_row, column = A_CI_col).value = str(df_part['CI Indicator (System Context)'][A_idx]) #'Part A CI Indicator
                    sheet.cell(row = current_row, column = B_CI_col).value = str(df_part['CI Indicator (System Context)'][A_idx]) #Part B CI Indicator
                    
                    APB = port_between(path[0], path[1], df_port, df_part, df_connector)
                    if not APB: 
                        A_PortName = 'No Port'
                        A_PortElmID = 'No Port, No ID'
                        A_Direction = 'No Port - Assumed Bi-Directional.'
                    else:
                        if type(df_port['Name'][APB[1]]) == str:
                            A_PortName = df_port['Name'][APB[1]]
                            A_PortElmID = df_port['Elm ID'][APB[1]]
                            A_Direction = str(df_port['Direction'][APB[1]])
                        else: 
                            A_PortName = 'Unnamed Port'
                            A_PortElmID = df_port['Elm ID'][APB[1]]
                            A_Direction = str(df_port['Direction'][APB[1]])
                            
                    sheet.cell(row = current_row, column = A_PortName_col ).value = A_PortName#'CI A - Port Name'
                    sheet.cell(row = current_row, column = A_PortElmID_col).value = A_PortElmID#'Port A - Element ID'
                    sheet.cell(row = current_row, column = A_Direction_col).value = A_Direction # CI A - Direction
                    
                    
                    Port_Types = list_or_set_to_string(pa[2])
                    sheet.cell(row = current_row, column = Port_Types_col).value = Port_Types#'Port Type(s)'
                    
                    Interface_Category = set()
                    for cat in pa[2]:
                        if cat in physicalIF: Interface_Category.add('Physical')
                        elif cat in logicalIF: Interface_Category.add('Logical')
                        elif cat in electricalIF: Interface_Category.add('Electrical')
                        elif cat in coolantIF: Interface_Category.add('Coolant')
                        else: Interface_Category.add('Interface in Path Not Categorized: ' + str(cat))
                    sheet.cell(row = current_row, column = Interface_Category_col).value = list_or_set_to_string(Interface_Category)
                    
                    B_Name = str(df_part['Name'][B_idx])
                    sheet.cell(row = current_row, column = B_Name_col).value = B_Name#'CI B - Name'
                    
                    B_Type = str(df_part['Type'][B_idx])
                    sheet.cell(row = current_row, column = B_Type_col).value = B_Type#'CI B - Type'
                    
                    B_Owner = str(df_part['Owner'][B_idx])
                    sheet.cell(row = current_row, column = B_Owner_col).value = B_Owner#'CI B - Owner'
                    
                    B_Context = list_or_set_to_string(B_context_list, 'No context', ' , ')
                    sheet.cell(row = current_row, column = B_Context_col).value = B_Context# 'CI B - Context'
                    
                    B_PhysicalJack = list_or_set_to_string(physical_jack(path[len(path)-3], path[len(path)-2], df_connector, df_part, df_port), 'No jack identified.')
                    sheet.cell(row = current_row, column = B_PhysicalJack_col).value = B_PhysicalJack #'CI B - Physical Jack'
                    
                    BPB = port_between(path[len(path)-2], path[len(path)-3], df_port, df_part, df_connector)
                    if not BPB: 
                        B_PortName = 'No Port'
                        B_PortElmID = 'No Port, No ID'
                        B_Direction = 'No Port - Assumed Bi-Directional.'
                    else:
                        if type(df_port['Name'][BPB[1]]) == str:
                            B_PortName = df_port['Name'][BPB[1]]
                            B_PortElmID = df_port['Elm ID'][BPB[1]]
                            B_Direction = str(df_port['Direction'][BPB[1]])
                        else: 
                            B_PortName = 'Unnamed Port'
                            B_PortElmID = df_port['Elm ID'][BPB[1]]
                            B_Direction = str(df_port['Direction'][BPB[1]])
                            
                    sheet.cell(row = current_row, column = B_PortName_col).value = B_PortName#'CI B - Port Name'                    
                    sheet.cell(row = current_row, column = B_PortElmID_col).value = B_PortElmID#'Port B - Element ID'
                    sheet.cell(row = current_row, column = B_Direction_col).value = B_Direction #CI -B - Direction
                    
                    sheet.cell(row = current_row, column = A_ElmID_col).value = path[0]
                    sheet.cell(row = current_row, column = B_ElmID_col).value = path[len(path) - 2]
                    sheet.cell(row = current_row, column = path_col).value = str(path)
                    
                    
                    
                    

                    # CI A Functional Area
                    A_FA = str(df_part['Functional Area (System Context)'][A_idx])
                    sheet.cell(row = current_row, column = A_FA_col).value = A_FA
                    
                    # CI B Functional Area
                    B_FA = str(df_part['Functional Area (System Context)'][B_idx])
                    sheet.cell(row = current_row, column = B_FA_col).value = B_FA
                    
                    # CI-A Name:CI-A Type (Assy / ... / Assy) from jack A through port A [gives/receives/gives & receives] [port type(s)]
                    sentence = A_Name + ':' + A_Type + ' (' + A_Context + ') ' + A_PhysicalJack + ' / ' + A_PortName + ' '
                    

                    # Gives (direction = out) / Receives (direction = in) / Gives & Receives (direction = inout) / Direction Undefined (direction = nan)
                    if A_Direction == 'out':
                        sentence += 'gives '
                    elif A_Direction == 'in':
                        sentence += 'receives '
                    elif A_Direction == 'inout':
                        sentence += 'gives & receives '
                    else:
                        sentence += 'unidentified direction '
                        
                    # Port Types
                    sentence += Port_Types + ' '
                    
                    # To (B direction = in) / From (B direction = out) / To & From (direction B = inout) / Direction Undefined (direction B = nan)
                    if B_Direction == 'in':
                        sentence += 'to '
                    elif B_Direction == 'out':
                        sentence += 'from ' 
                    elif B_Direction == 'inout':
                        sentence += 'to & from '
                    else:
                        sentence += 'unidenfied direction '
                        
                    # Name B : Type B (if unnamed, put in unnamed, if untyped, put in no type)
                    sentence += B_Name + ':' + B_Type + ' (' + B_Context + ') ' + B_Context + ') ' + B_PhysicalJack + ' / ' + B_PortName + '.'
                    
                    sheet.cell(row = current_row, column = sentence_col).value = sentence #'Sentence Description'
                    
                    path_direction = pa[5]
                    sheet.cell(row = current_row, column = path_direction_col).value = list_or_set_to_string(path_direction)

#------------------>#highlight cell if path is identified as having both an in port to in port or out port to out port
#                    if pa[6]:
#                        sheet.cell(row = current_row, column = path_direction_col).fill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type = 'solid')   #PatternFill(bgColor='FFC7CE', fill_type = 'solid')
                    
                    #ID the top level assembly that the path passes through.
                    sheet.cell(row = current_row, column = top_level_col).value = str(top_type)
                    sheet.cell(row = current_row, column = top_elmID_col).value = str(top_elmID)
                    
                    #ID the allocated wireharnesses
                    WHs = allocatedWireHarness(path, df_connector, df_part, df_port)
                    allocatedWH_names = []
                    if WHs:
                        for h in WHs:
                            [h_idx, h_type] = element_idx(h, df_port, df_part, df_connector)
                            if h_type == 'part': allocatedWH_names.append(str(df_part['Name'][h_idx]))

                    sheet.cell(row = current_row, column = allocatedWH_elmID_col).value = list_or_set_to_string(WHs)
                    sheet.cell(row = current_row, column = allocatedWH_name_col).value = list_or_set_to_string(allocatedWH_names, empty_message = 'No Allocated Wire Harness(es)', delim = ', ')
                    
                    tiers = getAssyTier(path[0], df_part)
                    tiers_len = len(tiers)
                    sheet.cell(row = current_row, column = A_Tier_col).value = tiers_len
                    sheet.cell(row = current_row, column = A_Tier1_col).value = tiers[0]
                    if tiers_len >= 2: sheet.cell(row = current_row, column = A_Tier2_col).value = tiers[1]
                    if tiers_len >= 3: sheet.cell(row = current_row, column = A_Tier3_col).value = tiers[2]
                    if tiers_len >= 4: sheet.cell(row = current_row, column = A_Tier4_col).value = tiers[3]
                    if tiers_len >= 5: sheet.cell(row = current_row, column = A_Tier5_col).value = tiers[4]
                    

                    Btiers = getAssyTier(path[len(path) - 2], df_part)
                    tiers_len = len(Btiers)
                    #sheet.cell(row = current_row, column = A_Tier_col).value = tiers_len
                    sheet.cell(row = current_row, column = B_Tier1_col).value = tiers[0]
                    if tiers_len >= 2: sheet.cell(row = current_row, column = B_Tier2_col).value = Btiers[1]
                    #if tiers_len >= 3: sheet.cell(row = current_row, column = A_Tier3_col).value = tiers[2]
                    #if tiers_len >= 4: sheet.cell(row = current_row, column = A_Tier4_col).value = tiers[3]
                    #if tiers_len >= 5: sheet.cell(row = current_row, column = A_Tier5_col).value = tiers[4]  
                                                                               
                    
                else: unwrittenPaths.append(path)
        except Exception as e:
            current_row = current_row + 1
            sheet.cell(row = current_row, column = 1).value = str(item)
            item = item + 1
            sheet.cell(row = current_row, column = 2).value = 'Error on index: ' + str(i)
            print(e)

################################################################################
# Adjust column widths so that it is readable
columnletters = ('', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ')
sheet.column_dimensions[columnletters[A_Type_col]].width = 40
sheet.column_dimensions[columnletters[A_Name_col]].width = 35.43
sheet.column_dimensions[columnletters[B_Name_col]].width = 35.43
sheet.column_dimensions[columnletters[A_PortName_col]].width = 37.14
sheet.column_dimensions[columnletters[B_PortName_col]].width = 37.14
sheet.column_dimensions[columnletters[A_PhysicalJack_col]].width = 37.14
sheet.column_dimensions[columnletters[B_PhysicalJack_col]].width = 37.14
sheet.column_dimensions[columnletters[allocatedWH_name_col]].width = 28.43
sheet.column_dimensions[columnletters[B_Type_col]].width = 40
sheet.column_dimensions[columnletters[A_Owner_col]].width = 36.29
sheet.column_dimensions[columnletters[B_Owner_col]].width = 36.29
sheet.column_dimensions[columnletters[A_FA_col]].width = 14
sheet.column_dimensions[columnletters[B_FA_col]].width = 14
sheet.column_dimensions[columnletters[Interface_Category_col]].width = 18.14
sheet.column_dimensions[columnletters[Port_Types_col]].width = 19.71
sheet.column_dimensions[columnletters[A_Kind_col]].width = 28.86
sheet.column_dimensions[columnletters[A_Direction_col]].width = 16.86
sheet.column_dimensions[columnletters[B_Direction_col]].width = 16.86
sheet.column_dimensions[columnletters[sentence_col]].width = 40
sheet.column_dimensions[columnletters[A_Tier_col]].width = 5
sheet.column_dimensions[columnletters[A_Tier2_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier3_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier4_col]].width = 36.29
sheet.column_dimensions[columnletters[A_Tier5_col]].width = 36.29
sheet.column_dimensions[columnletters[item_col]].width = 19.71



# used to auto generate column widths, but changed in favor to above manually setting them to keep compatct
#for col in sheet.columns:
#     max_length = 0
#     column = col[0].column # Get the column name
#     for i, cell in enumerate(col):
#         try: # Necessary to avoid error on empty cells
#             if (i not in [0, 1, 3]) and len(str(cell.value)) > max_length:
#                 max_length = len(cell.value)
#         except:
#             pass
#     if max_length < 50: adjusted_width = (max_length)
#     else: adjusted_width = int(round(max_length*.8))
#     sheet.column_dimensions[column].width = adjusted_width


# Hide MagicDraw Specific Info
if hidemagicdraw:
    sheet.column_dimensions[openpyxl.utils.get_column_letter(A_ElmID_col)].hidden = True # hide columns with element IDs
    sheet.column_dimensions[openpyxl.utils.get_column_letter(B_ElmID_col)].hidden = True
    sheet.column_dimensions[openpyxl.utils.get_column_letter(path_col)].hidden = True
    sheet.column_dimensions[openpyxl.utils.get_column_letter(A_PortElmID_col)].hidden = True    
    sheet.column_dimensions[openpyxl.utils.get_column_letter(B_PortElmID_col)].hidden = True
    sheet.column_dimensions[openpyxl.utils.get_column_letter(path_direction_col)].hidden = True


################################################################################
# Write Sheet with additional information
wb.create_sheet(index=2, title='Column Definitions')
cd_ws = wb.get_sheet_by_name('Column Definitions')

cd_ws.cell(row = 1, column = 1).value = 'Column Header'
cd_ws.cell(row = 1, column = 2).value = 'Definition'
cd_ws.cell(row = 1, column = 1).font = bold14
cd_ws.cell(row = 1, column = 2).font = bold14

cd_ws.cell(row = 2, column = 1).value = 'Part A Kind'
cd_ws.cell(row = 2, column = 2).value = "This is Part A's categorization as an item of software, wire harness, or hardware or specific hardware kind."

cd_ws.cell(row = 3, column = 1).value = 'Wire Harness'
cd_ws.cell(row = 3, column = 2).value = 'This is the wire harness(es) to which this connection path is allocated.  For a given wire harness, you can filter to find all sets of connections that must be included in that wire harness.'

cd_ws.cell(row = 4, column = 1).value = 'Part A or B Functional Area'
cd_ws.cell(row = 4, column = 2).value = 'This is the functional area to which Part A or B belong, e.g., EXO, PWR, etc.'

cd_ws.cell(row = 5, column = 1).value = 'Part A or B Name'
cd_ws.cell(row = 5, column = 2).value = 'This is the name for the specific instance that is Part A or Part B, e.g., the left elbow UEX actuator.'

cd_ws.cell(row = 6, column = 1).value = 'Part A or B CI'
cd_ws.cell(row = 6, column = 2).value = 'This is the type of CI that is Part A or Part B, e.g. a UEX actuator'

cd_ws.cell(row = 7, column = 1).value = 'Part A or B Assembly'
cd_ws.cell(row = 7, column = 2).value = 'This is the next higher assembly to which Part A or Part B belongs, e.g. the left elbow sensing and actuation group.'

cd_ws.cell(row = 8, column = 1).value = 'Part A or B Context'
cd_ws.cell(row = 8, column = 2).value = 'This show the set of all assemblies the path between part A and part B travel through.'

cd_ws.cell(row = 9, column = 1).value = 'Part A or B Physical Jack'
cd_ws.cell(row = 9, column = 2).value = 'This is the physical jack where this connection occurs on part A or B, e.g., Hirose connector xyz'

cd_ws.cell(row = 10, column = 1).value = 'Part A or B Functional Interface Type'
cd_ws.cell(row = 10, column = 2).value = 'This functional or logical type of port that the above jack is, e.g. a high voltage'

cd_ws.cell(row = 11, column = 1).value = 'Part A or B Jack Flow Direction'
cd_ws.cell(row = 11, column = 2).value = 'This is the direction that information or energy flows from the jack, e.g., in means that the jack receives information, out means it gives information, and inout means it is bidirectional'

cd_ws.cell(row = 12, column = 1).value = 'Connection Category'
cd_ws.cell(row = 12, column = 2).value = 'This is a categorization of connection types as electrical, logical, or physical.'

cd_ws.cell(row = 13, column = 1).value = 'Connection Type(s)'
cd_ws.cell(row = 13, column = 2).value = 'This is the set of port types through which the path travels.'

cd_ws.cell(row = 14, column = 1).value = 'Connection #'
cd_ws.cell(row = 14, column = 2).value = 'This is an arbitrary number assigned when developing this list; it can change based on input order.  DO NOT use it for tracking purposes.  If we need unique ID numbers for these connections, please let Steve Gillespie know and he can come up with a unique identifier.'

cd_ws.cell(row = 15, column = 1).value = 'Part A Assembly Tier X'
cd_ws.cell(row = 15, column = 2).value = "This is the Xth tier that part A resides in.  Note that Tier 1 should always be 'FULL TALOS Assembly'" 

cd_ws.cell(row = 16, column = 1).value = 'Part A Assembly Tier'
cd_ws.cell(row = 16, column = 2).value = "This is the lowest numbered tier that part A resides in" 

cd_ws.column_dimensions['A'].width = 30
cd_ws.column_dimensions['B'].width = 200
     
# Finally save the output to the desired directory with the name:
# CI_to_CI_Analysis_MODEL#.xlsx
os.chdir(output_file_directory)
wb.save('CI_to_CI_Analysis_' + str(model_version) + '.xlsx')


if True:
    import pandas as pd
    df_connections = pd.read_excel('CI_to_CI_Analysis_' + str(model_version) + '.xlsx', sheetname = 'Connections', skiprows = 4)
    wb.create_sheet(index=2, title='CI Ports')
    cip_ws = wb.get_sheet_by_name('CI Ports')
    
    cip_ws.cell(row = 1, column = 1).value = 'Part A CI'
    cip_ws.cell(row = 1, column = 2).value = 'Part A Functional Interface Type'
    cip_ws.cell(row = 1, column = 3).value = 'Set of Ports on CIs Where This is Attached'
    cip_ws.cell(row = 1, column = 1).font = bold14
    cip_ws.cell(row = 1, column = 2).font = bold14
    cip_ws.cell(row = 1, column = 3).font = bold14
        
    current_row = 2            
    previouslyCheckedCIPorts = set()
    for i in np.arange(len(df_connections)):
        if type(df_connections['Part A CI'][i]) != str: continue
        CI_Port = (df_connections['Part A CI'][i], df_connections['Part A Functional Interface Type'][i])
        if CI_Port in previouslyCheckedCIPorts: continue
        previouslyCheckedCIPorts.add(CI_Port)
        
        # find all indices in the connections table who have the given Part A CI and Part A Port
        idx_to_check = df_connections.loc[(df_connections['Part A CI'] == CI_Port[0]) & (df_connections['Part A Functional Interface Type'] == CI_Port[1])].index.tolist()
        set_of_ports = set()
        for idx in idx_to_check: set_of_ports.add((df_connections['Part B CI'][idx], df_connections['Part B Functional Interface Type'][idx]))
        
        cip_ws.cell(row = current_row, column = 1).value = df_connections['Part A CI'][i]
        cip_ws.cell(row = current_row, column = 2).value = df_connections['Part A Functional Interface Type'][i]        
        CIPs = ''
        for sop in set_of_ports: CIPs += (sop[1] + ' on ' + sop[0] + '\n')
        cip_ws.cell(row = current_row, column = 3).value = CIPs[0:len(CIPs) - 1]
        cip_ws.cell(row = current_row, column = 3).alignment = openpyxl.styles.Alignment(wrapText = True)
        current_row += 1   
    
    cip_ws.column_dimensions['A'].width = 32
    cip_ws.column_dimensions['B'].width = 32
    cip_ws.column_dimensions['C'].width = 100
    
    wb.save('CI_to_CI_Analysis_' + str(model_version) + '.xlsx')

    
####################
####################
# BUILD CONNECTION ANALYSIS TABLE (Tier 2)

connections_df = XLSX_to_DF(output_file_directory, 'CI_to_CI_Analysis_' + str(model_version) + '.xlsx', 'Connections', 4)
error_idx = []

checked_connections = []
results_df = pd.DataFrame(columns = ['Logical_Internal', 'Physical_Internal', 'Electrical_Internal', 'Logical_External', 'Physical_External', 'Electrical_External'])

IF_risk = []

for i in np.arange(1, len(connections_df)):
    #print(i)
    #don't double count connections
    if connections_df['Element ID Top Connector'][i] in checked_connections: continue
    checked_connections.append(connections_df['Element ID Top Connector'][i])
    
    if connections_df['Part A Functional Area'][i] != connections_df['Part B Functional Area'][i]:
        # Part A Name, Part A FA, Part A Elm ID, Part B Name, Part B FA, Part BA Elm ID, Top Connector Elm ID
        
        IF_risk.append([connections_df['Part A Name'][i], connections_df['Part A Functional Area'][i], 
        connections_df['Part A MagicDraw ID'][i], connections_df['Part B Name'][i], 
        connections_df['Part B Functional Area'][i], connections_df['Part B MagicDraw ID'][i],
        connections_df['Element ID Top Connector'][i]])
        
    ## count connections inside and outside of Tier 2 assemblies
    if type(connections_df['Part A Assembly Tier 2'][i]) != str or type(connections_df['Part B Assembly Tier 2'][i]) != str:
        print('NaN Tier 2s at index: ' + str(i))
        continue
    
    
    if connections_df['Part A Assembly Tier 2'][i] == connections_df['Part B Assembly Tier 2'][i]:
        if connections_df['Part A Assembly Tier 2'][i] not in results_df.index: results_df.loc[connections_df['Part A Assembly Tier 2'][i]] = np.zeros(len(results_df.columns))
        
        if connections_df['Connection Category'][i] == 'Logical':  results_df['Logical_Internal'][connections_df['Part A Assembly Tier 2'][i]] += 1
        elif connections_df['Connection Category'][i] == 'Electrical':  results_df['Electrical_Internal'][connections_df['Part A Assembly Tier 2'][i]] += 1
        elif connections_df['Connection Category'][i] == 'Physical':  results_df['Physical_Internal'][connections_df['Part A Assembly Tier 2'][i]] += 1
    
    else:
        if connections_df['Part A Assembly Tier 2'][i] not in results_df.index: results_df.loc[connections_df['Part A Assembly Tier 2'][i]] = np.zeros(len(results_df.columns))
        if connections_df['Part B Assembly Tier 2'][i] not in results_df.index: results_df.loc[connections_df['Part B Assembly Tier 2'][i]] = np.zeros(len(results_df.columns))
        
        if connections_df['Connection Category'][i] == 'Logical':  
            results_df['Logical_External'][connections_df['Part A Assembly Tier 2'][i]] += 1
            results_df['Logical_External'][connections_df['Part B Assembly Tier 2'][i]] += 1
        elif connections_df['Connection Category'][i] == 'Electrical':  
            results_df['Electrical_External'][connections_df['Part A Assembly Tier 2'][i]] += 1
            results_df['Electrical_External'][connections_df['Part B Assembly Tier 2'][i]] += 1
        elif connections_df['Connection Category'][i] == 'Physical':  
            results_df['Physical_External'][connections_df['Part A Assembly Tier 2'][i]] += 1
            results_df['Physical_External'][connections_df['Part B Assembly Tier 2'][i]] += 1



wb.create_sheet(index=3, title='Connection Count')
cc_ws = wb.get_sheet_by_name('Connection Count')


from openpyxl.utils.dataframe import dataframe_to_rows
rows = dataframe_to_rows(results_df)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         cc_ws.cell(row=r_idx, column=c_idx, value=value)

wb.save('CI_to_CI_Analysis_' + str(model_version) + '.xlsx')


# Enumerate IF Risk (by FA)
wb.create_sheet(index=4, title='Cross FA Interfaces')
cfi_ws = wb.get_sheet_by_name('Cross FA Interfaces')

cfi_ws.cell(row = 1, column = 1).value = 'Part A Name'
cfi_ws.cell(row = 1, column = 2).value = 'Part A FA'
cfi_ws.cell(row = 1, column = 3).value = 'Part A ID'
cfi_ws.cell(row = 1, column = 4).value = 'Part B Name'
cfi_ws.cell(row = 1, column = 5).value = 'Part B FA'
cfi_ws.cell(row = 1, column = 6).value = 'Part B ID'
cfi_ws.cell(row = 1, column = 7).value = 'Assigned Connector'

r = 2
for risk in IF_risk:
    cfi_ws.cell(row = r, column = 1).value = risk[0]
    cfi_ws.cell(row = r, column = 2).value = risk[1]
    cfi_ws.cell(row = r, column = 3).value = risk[2]
    cfi_ws.cell(row = r, column = 4).value = risk[3]
    cfi_ws.cell(row = r, column = 5).value = risk[4]
    cfi_ws.cell(row = r, column = 6).value = risk[5]
    cfi_ws.cell(row = r, column = 7).value = risk[6]
    r += 1

wb.save('CI_to_CI_Analysis_' + str(model_version) + '.xlsx')
